########################################################################
########################        Libraries       ########################

#   - Help can be in the tools.json; but remember to update Discord, Telegram, and Line ~ Maybe not really relevant and might interfere with her response on what tools she have
#   - Web is Scraping for Articles and Youtube Music and Spotify ( Passing to ChatGPT for Summary is done in the respective Scripts )
#   - SearchGPT ( Hope we get on the list )
#   - Async is only used for Discord
#   - Telegram and Line have their individual Handlers
#   - Done Slack 
#   - Done Notes ( Saved to User_ID )  [ Save in Server for Now; Save in MongoDB next time ]
#   - Fixed the Character Encoding for the Logger { Windows doesn't use fucking UTF-8 Codec } ( Highly Annoying !!! )

# You can Build Draw Here to send Back the Image Reply together with UPDATED VALUES * Can do the Same with Chat
#   -   Return with multiple Varibles, that way you can update back to the Global of the Other Class ( This is Hard to Do; You have to radically change the Structures )

# User Differentiated [ Build each Layer from Class instead of Json ]

# Iterate the Users' Filename ( Do when you have a bigger Userbase * Use a bigger Database [ MongoDB ])  This is somewhat done when we store the binaries of their sent files in MongoDB

# Lyrics ( All the Libraries I use don't really work anymore ) [ What is the best site / app for this ? ]  Use Web Scraper ?  * Just give the Link instead of scraping the whole site ?

# Build GeoLocation ~ GPS Python [ Google Maps ] ( Apple Maps )

# Audio Voice Transcription { OpenAI }

# OpenAI has finally released the Web Search API ; Now we can get replies for real time information and services
# Grok has best voice
# Claude is catching up

# Create Config Template

# Memory History should be Contextualized

# Global Memory [ Unique User IDs ]

import os
import io
import re
import sys
import time
import pytz
import json
import math
import copy
import base64
import random
import string
import psutil
import logging
import asyncio
import requests
import platform

import yt_dlp
import wikipediaapi
import urllib.parse
import pandas as pd
from gtts import gTTS
from ytmusicapi import YTMusic
from youtube_search import YoutubeSearch
from discord import FFmpegPCMAudio, PCMVolumeTransformer
from linebot.models import TextSendMessage, ImageSendMessage

from openai import OpenAI
from langdetect import detect
from pydub import AudioSegment
from shelly import ShellyManager, ShellyError
from datetime import datetime, timedelta, timezone

from pprint import pprint
from bs4 import BeautifulSoup
from newsapi import NewsApiClient
from openpyxl import Workbook, load_workbook
from concurrent.futures import ThreadPoolExecutor
from urllib.parse import urlparse, parse_qsl, urlencode, urlunparse, quote_plus, quote

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.firefox import GeckoDriverManager
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.firefox.service import Service as FirefoxService
import selenium.common.exceptions as SeleniumException

########################################################################
##########################        Codec       ##########################

# Switch Windows console to UTF-8 code page
if os.name == "nt": os.system("chcp 65001 > nul")

try:
    # Python 3.7+: reconfigure stdout/stderr to UTF-8
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    # Fallback (e.g. older Python or non-TextIO stdout)
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

########################################################################
##########################        Tools       ##########################

# Media
AudioTool = {
    "type": "function",
    "function": {
        "name": "Audio",
        "description": "Turn ON/OFF audio listening features (voice message transcription). Use this when the user asks to enable/disable audio detection or listening.",
        "parameters": {
            "type": "object",
            "properties": {
                "state": {
                    "type": "string",
                    "enum": ["on", "off", "toggle", "status"],
                    "description": "Set or toggle audio listening. Use 'status' to query current state."
                }
            },
            "required": ["state"]
        }
    }
}
SpeechTool = {
    "type": "function",
    "function": {
        "name": "Speak",
        "description": "Turn ON/OFF voice replies (text-to-speech). Use this when the user asks you to speak or stop speaking.",
        "parameters": {
            "type": "object",
            "properties": {
                "state": {
                    "type": "string",
                    "enum": ["on", "off", "toggle", "status"],
                    "description": "Set or toggle speaking (TTS). Use 'status' to query current state."
                }
            },
            "required": ["state"]
        }
    }
}
VisionTool = {
    "type": "function",
    "function": {
        "name": "Vision",
        "description": "Turn ON/OFF image understanding features (describe images/videos). Use this when the user asks to enable/disable image description.",
        "parameters": {
            "type": "object",
            "properties": {
                "state": {
                    "type": "string",
                    "enum": ["on", "off", "toggle", "status"],
                    "description": "Set or toggle vision features. Use 'status' to query current state."
                }
            },
            "required": ["state"]
        }
    }
}
DrawTool = {
    "type": "function",
    "function": {
        "name": "Draw",
        "description": "Generate an image based on a text prompt",
        "parameters": {
            "type": "object",
            "properties": {
                "prompt": {
                    "type": "string",
                    "description": "The text prompt to generate an image for"
                }
            },
            "required": ["prompt"]
        }
    }
}
WebTool = {
    "type": "function",
    "function": {
        "name": "Web",
        "description": "Performs a web search and returns information sourced from the internet using the Web Search Preview Model.",
        "parameters": {
            "type": "object",
            "properties": {
                "query": {
                    "type": "string",
                    "description": "The search query to find information on the web"
                }
            },
            "required": ["query"]
        }
    }
}

# Auxiliary 
WeatherTool = {
    "type": "function",
    "function": {
        "name": "Weather",
        "description": "Provide weather details based on a city",
        "parameters": {
            "type": "object",
            "properties": {
                "location": {
                    "type": "string",
                    "description": "The city to get the weather for"
                }
            },
            "required": [
                "location"
            ]
        }
    }
}
MusicTool = {
    "type": "function",
    "function": {
        "name": "Music",
        "description": "Play music based on a song",
        "parameters": {
            "type": "object",
            "properties": {
                "song": {
                    "type": "string",
                    "description": "The song to search music for"
                }
            },
            "required": [
                "song"
            ]
        }
    }
}
YoutubeTool = {
    "type": "function",
    "function": {
        "name": "Youtube",
        "description": "Play a video on YouTube based on a query",
        "parameters": {
            "type": "object",
            "properties": {
                "video": {
                    "type": "string",
                    "description": "The search query for the YouTube video"
                }
            },
            "required": [
                "video"
            ]
        }
    }
}
NoteTool =  {
    "type": "function",
    "function": {
        "name": "Note",
        "description": "Write a new reminder",
        "parameters": {
            "type": "object",
            "properties": {
                "reminder": {
                    "type": "string",
                    "description": "The reminder note you want to add"
                }
            },
            "required": [
                "reminder"
            ]
        }
    }
}
PeekTool = {
    "type": "function",
    "function": {
        "name": "Peek",
        "description": "Read or Clear existing reminders",
        "parameters": {
            "type": "object",
            "properties": {
                "command": {
                    "type": "string",
                    "description": "Return 'read' or 'clear'"
                }
            },
            "required": [
                "command"
            ]
        }
    }
}
CurrencyTool = {
    "type": "function",
    "function": {
        "name": "Currency",
        "description": "Converts an amount from one currency to another",
        "parameters": {
            "type": "object",
            "properties": {
                "from": {
                    "type": "string",
                    "description": "The currency code to convert from (e.g., 'USD')"
                },
                "to": {
                    "type": "string",
                    "description": "The currency code to convert to (e.g., 'EUR')"
                },
                "amount": {
                    "type": "number",
                    "description": "The amount to be converted"
                }
            },
            "required": [
                "from_currency",
                "to_currency",
                "amount"
            ]
        }
    }
}
CryptoTool = {
    "type": "function",
    "function": {
        "name": "Crypto",
        "description": "Converts an amount from one crypto to another",
        "parameters": {
            "type": "object",
            "properties": {
                "from": {
                    "type": "string",
                    "description": "The crypto code to convert from (e.g., 'BTC')"
                },
                "to": {
                    "type": "string",
                    "description": "The crypto code to convert to (e.g., 'ETH')"
                },
                "amount": {
                    "type": "number",
                    "description": "The amount to be converted"
                }
            },
            "required": [
                "from_currency",
                "to_currency",
                "amount"
            ]
        }
    }
}
TimeTool = {
    "type": "function",
    "function": {
        "name": "Time",
        "description": "Convert time between two time zones",
        "parameters": {
            "type": "object",
            "properties": {
                "from": {
                    "type": "string",
                    "description": "The originating time zone in TZ database format (e.g., 'Asia/Bangkok')"
                },
                "to": {
                    "type": "string",
                    "description": "The target time zone in TZ database format (e.g., 'Asia/Kolkata')"
                },
                "time": {
                    "type": "string",
                    "description": "The time to convert in format 'HH:MM AM/PM'"
                }
            },
            "required": [
                "from",
                "to",
                "time"
            ]
        }
    }
}
MapsTool = {
    "type": "function",
    "function": {
        "name": "Maps",
        "description": "Generate a Google Maps URL from natural language. Supports directions ('A -> B [driving|walking|bicycling|transit]' or 'from A to B [mode]') and search ('<keyword> near/nearby <place>' or a plain place). Returns a single ready-to-open URL.",
        "parameters": {
            "type": "object",
            "properties": {
                "text": {
                "type": "string",
                "description": "User input, e.g. 'from Bang Na to ICONSIAM transit', 'Thai food near Ekkamai', 'pizza nearby Thonglor', or 'ICONSIAM'."
                }
            },
            "required": [
                "text"
            ]
        }
    }
}

# Embedding
VectorTool = {
    "type": "function",
    "function": {
        "name": "Vector",
        "description": "Search the local vector store for answers from uploaded documents",
        "parameters": {
        "type": "object",
        "properties": {
            "query": {
            "type": "string",
            "description": "The user’s question to search against the vector embeddings"
            }
        },
        "required": ["query"]
        }
    }
}
ProfileTool = {
    "type": "function",
    "function": {
        "name": "Profile",
        "description": "Look up a value from the user's vector/profile store using a dot path (e.g., 'passport.expiry', 'identity.full_name', 'contacts[0].email'). Use this whenever the user asks about their details.",
        "parameters": {
        "type": "object",
        "properties": {
            "field": {
                "type": "string",
                "description": "Dot path into the vector/profile JSON, e.g. 'loyalty.krisflyer' or 'passport.expiry'"
                },
            "default": {
                "type": "string",
                "description": "Optional fallback value if the field is missing",
                "default": ""
                }
            },
        "required": ["field"]
        }
    }
}
ListingTool = {
    "type": "function",
    "function": {
        "name": "Listings",
        "description": "Search the real-estate listing demo and return the best matching properties.",
        "parameters": {
        "type": "object",
        "properties": {
            "query": {
                "type": "string",
                "description": "Natural language description of what the buyer or investor wants, e.g. 'sea-view loft in Limassol under €400k with 2 bedrooms and >5.8% ROI'"
            },
            "max_results": {
                "type": "integer",
                "description": "Maximum number of listings to return",
                "default": 3
            }
        },
        "required": ["query"]
        }
    }
}

# Home
ShellyTool = {
    "type": "function",
    "function": {
        "name": "Shelly",
        "description": (
            "Control Shelly Plus 2PM relays/roller. "
            "Use on/off/toggle for switching, status to check on/off. "
            "IMPORTANT: Use power ONLY when the user explicitly asks about watts/energy usage."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "device": {
                    "type": "string",
                    "enum": [
                        "library",
                        "restroom",
                        "bedroom",
                        "balcony",
                        "closet",
                        "bathroom",
                        "sky",
                        "deck",
                        "loft"
                    ],
                    "description": "Logical Shelly device name."
                },
                "channel": {
                    "type": "integer",
                    "description": "Relay channel index on Shelly Plus 2PM (0 or 1)."
                },
                "action": {
                    "type": "string",
                    "description": (
                        "Action for relay/roller. "
                        "Choose 'power' ONLY for explicit power/energy questions (watts/kWh)."
                    ),
                    "enum": [
                        "on",
                        "off",
                        "toggle",
                        "status",
                        "cover_open",
                        "cover_close",
                        "cover_stop",
                        "power"
                    ]
                },
                "duration": {
                    "type": "integer",
                    "description": "Optional seconds for timed cover open/close (roller mode)."
                },
                "mode": {
                    "type": "string",
                    "description": "Response compactness. Use 'compact' unless debugging.",
                    "enum": ["compact", "debug"],
                    "default": "compact"
                }
            },
            "required": ["device", "channel", "action"]
        }
    }
}

########################################################################
########################        Interface       ########################

# Interface
class Interface:
    ########################################################################
    ##################                Init                ##################

    def __init__(self, settings):
        self.Settings         = settings
        self.executor         = ThreadPoolExecutor(max_workers=10)
        self.Emoji_Pattern    = re.compile(r'[\U0001F600-\U0001F64F]')
        self.Emoji_Extended   = re.compile(
                                        r'['
                                        r'\U0001F600-\U0001F64F'  # smileys
                                        r'\U0001F300-\U0001F5FF'  # symbols & pictographs
                                        r'\U0001F680-\U0001F6FF'  # transport & map
                                        r'\U0001F700-\U0001F77F'  # alchemical
                                        r'\U0001F900-\U0001F9FF'  # supplemental symbols
                                        r'\U0001FA70-\U0001FAFF'  # newer symbols
                                        r'\U00002600-\U000026FF'  # misc symbols
                                        r'\U00002700-\U000027BF'  # dingbats
                                        r']')
        self.URL_Pattern      = r'http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\\(\\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+'
        self.fay_Pattern      = r'(?:fay|faye|(?<!\w)เฟ(?!\w)|เฟย|เฟย์)'  # fay_pattern = r'\bfay\b|\bfaye\b|\bเฟ\b|\bเฟย\b|\bเฟย์\b' * ink_Pattern = r'\bink\b|\bอิ้ง\b|\bอิ้งค\b|\bอิงค์\b'
        self.ink_Pattern      = r'(?:ink|(?<!\w)อิ้ง(?!\w)|อิ้งค|อิงค์)'  
        self.Email_Pattern    = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b'

        self.ShellyDevices    = self.Settings.ShellyDevices         
        self.ShellyAliases    = self.Settings.ShellyAliases
        self.ShellyManager    = ShellyManager(self.ShellyDevices)

    def Reply(self, *args):
        pass

    def Save(self, *args):
        pass

    def Listener(self, *args):
        pass

    def Name(self, name):
        # Sanitize the name to match OpenAI's expected pattern ( Line has their own set of Emojis that doesn't adhere to normal character encoding for the computer )
        name = re.sub(r"[^a-zA-Z0-9_-]", "", name)      # Replace Spaces
        #name = re.sub(r"[^a-zA-Z0-9 _-]", "", name)    # Allows Spaces ( Does OpenAI Name Argument allow Spaces ? It DOESN'T )
        return name or "UnknownUser"

    def Logging(self, info="info", error="error"):
        # Logging
        logger = logging.getLogger(__name__)
        logger.setLevel(logging.DEBUG)  # Capture all levels of messages

        # Handler for INFO and lower‐level logs, with UTF-8 encoding
        info_handler = logging.FileHandler(f"{info}.log", encoding="utf-8")
        info_format  = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")
        info_handler.setLevel(logging.INFO)
        info_handler.setFormatter(info_format)
        logger.addHandler(info_handler)

        # Handler for ERROR and higher‐level logs, with UTF-8 encoding
        error_handler = logging.FileHandler(f"{error}.log", encoding="utf-8")
        error_format  = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")
        error_handler.setLevel(logging.ERROR)
        error_handler.setFormatter(error_format)
        logger.addHandler(error_handler)
        return logger

    def Terminal(self, logger, text, type="info"):
        if type == "info":
            print(text)
            logger.info(text)
        else:
            print(f"Error: {text}")
            logger.error(text)
        return

    def Sys(self, *args):
        processor   = f"Processor: {platform.processor()}"
        memory_info = psutil.virtual_memory()
        memory      = f"Memory: {round(memory_info.total/1_073_741_824, 2)}GB total, {round(memory_info.available/1_073_741_824, 2)}GB available, {memory_info.percent}% used"
        disk_info   = psutil.disk_usage('/')
        disk        = f"Disk Space: {round(disk_info.total/1_073_741_824, 2)}GB total, {round(disk_info.free/1_073_741_824, 2)}GB free, {disk_info.percent}% used"
        output      = f"{processor}\n{memory}\n{disk}"
        return output

    def Help(self, name, *args):
        menu        = f"""Help Menu:

        1.  /token <value>: 
        -   Description: Set the token value from 1 to 4096.
        -   Example: /token 2500
        -   Tip: Use this to control the Output
            ( How many Words you want {name} to Reply )
        -   Note: If {name} cuts off, you can simply say "Continue"

        2.  /model <model_name>: 
        -   Description: Sets the model. Models: GPT4M, GPT4O, GPT4T
        -   Example: /model GPT4M
        -   Tip: Use this if you are experiencing errors with your model
        -   Note: This will Clear the Memory as well

        3.  /refresh: 
        -   Description: Refreshes the Memory immediately.
        -   Example: /refresh
        -   Tip: Use this to Refresh the Memory 
            If you want to talk about something new with a fresh memory

        4.  /drawsize <size>: 
        -   Description: Sizes: Normal, Landscape, Portrait.
        -   Example: /drawsize Landscape
                        
        5.  /draw <prompt>
        -   Description: Generates an image based on the text prompt.
        -   Example: /draw a beautiful sunset over the mountains
        -   Tip: Use /draw refresh 
            If you want to Draw something new with a fresh memory
        -   Note: Draw has a different Memory than the Chat Memory

        6.  /vision
        -   Description: Turns on Vision
        -   Example: /vision
        -   Tip: Use this to Activate the Vision Feature 
            
        7.  /speak
        -   Description: Turns on Voice
        -   Example: /speak
        -   Tip: Use this to Activate the Voice Feature

        8.  /listen
        -   Description: Turns on Listening
        -   Example: /listen
        -   Tip: Use this to Activate the Listen Feature

        9.  /web
        -   Description: Searches the Web for Infomation
        -   Example: /web {{ query }}
        -   Tip: Use this if you want your information to be cited by sources 
        """
        return menu

    def MarkupStrip(self, text):
        if not isinstance(text, str):   return text

        # **bold** and __bold__
        text = re.sub(r'\*\*(.+?)\*\*', r'\1', text, flags=re.DOTALL)
        text = re.sub(r'__(.+?)__',   r'\1', text, flags=re.DOTALL)

        # *italics* and _italics_
        # Guard against list bullets: require no whitespace immediately after the opening marker
        # and no whitespace immediately before the closing marker.
        text = re.sub(r'(?<!\w)\*(?!\s)(.+?)(?<!\s)\*(?!\w)', r'\1', text, flags=re.DOTALL)
        text = re.sub(r'(?<!\w)_(?!\s)(.+?)(?<!\s)_(?!\w)',   r'\1', text, flags=re.DOTALL)
        return text

    def EmojiCheck(self, text) :    return bool(self.Emoji_Pattern.fullmatch((text or "").strip()))

    def TimeStamp(self):            return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ########################################################################
    ##################                Core                ##################

    def Chat(self, name, user_id, chat_id, platform, client, model, memory, token, temperature, tools, prompt, logger, mediaDict):
        # ChatCompletion
        def get_response(model, memory, token, temperature, tools, bTools):  
            """
            ChatCompletions wrapper.
            - Uses your internal memory structure:
              { gptRole, gptName, gptContent }
            - Sends it to client.chat.completions.create(...)
            - Supports temperature + tools + tool_calls
            """
            try:
                kwargs = {
                    "model": model,            # e.g. "gpt-5", "gpt-4o", etc. (real model id)
                    "messages": memory,
                    "temperature": temperature
                }

                # Temperature { GPT-5 might not support } [ Controlled from the outside ]
                # kwargs["temperature"] = random.uniform(0.5, 1.2)

                # Tokens: 
                if token:   
                    if model == "gpt-5":    kwargs["max_completion_tokens"] = token # GPT5 is broken for now ! TRY IT
                    else:                   kwargs["max_tokens"] = token
                        
                # Tools: 
                if tools:   
                    kwargs["tools"]         = tools
                    kwargs["tool_choice"]   = "auto" if bTools else "none"  # "required" forces a tool Call
                
                # Convert image format for Claude [ Use Claude SDK ]
                if "claude" in model.lower():
                    memory = copy.deepcopy(memory)
                    for msg in memory:
                        content = msg.get(self.Settings.gptContent)
                        if isinstance(content, list):
                            for i, block in enumerate(content):
                                if isinstance(block, dict) and block.get("type") == "image_url":
                                    url = block["image_url"]["url"]
                                    if url.startswith("data:image/"):
                                        header, b64 = url.split(",", 1)
                                        media_type = header.split(":")[1].split(";")[0]
                                        content[i] = {
                                            "type": "image",
                                            "source": {"type": "base64", "media_type": media_type, "data": b64}
                                        }
                                    else:
                                        content[i] = {
                                            "type": "image",
                                            "source": {"type": "url", "url": url}
                                        }
                    kwargs["messages"] = memory
                
                # Verbosity and Web Search

                # API
                result          = client.chat.completions.create(**kwargs)
                response        = result.choices[0].message.content or ""
                function        = result.choices[0].message.tool_calls
                tokens          = result.usage.total_tokens
                temperature     = temperature
                return response, function
            except Exception as e:
                error = f"Error: {e}\n"
                self.Terminal(logger, error, "error")
                return error, None
        
        # ResponseAPI
        ''' 
        def get_response(memory):
            try:
                # Convert Internal Memory → Responses API messages
                api_input = []
                for m in memory or []:
                    role    = m.get(self.Settings.gptRole)    or m.get("role")    or "user"
                    content = m.get(self.Settings.gptContent) or m.get("content") or ""

                    api_input.append({
                        "role": role,
                        "content": str(content),
                    })

                kwargs = {
                    "model": model,
                    "input": api_input,   # <— array of messages, no types [ "temperature": random.uniform(0.5, 1.2) ] 
                }

                if token:   kwargs["max_output_tokens"] = token
                if tools:   kwargs["tools"] = self._normalize_tools_for_responses(tools)
                
                # Response
                result = client.responses.create(**kwargs)  

                print(result)
                reply      = getattr(result, "output_text", "") or ""

                if not reply and hasattr(result, "output"):
                    chunks = []
                    for out in result.output:
                        # out.type might be "message", "output_text", etc.
                        content = getattr(out, "content", None) or []
                        for part in content:
                            if isinstance(part, dict) and part.get("type") in ("output_text", "output_text_delta"):
                                chunks.append(part.get("text", ""))
                    reply = "\n".join(filter(None, chunks))
                
                tool_calls = getattr(result, "output_tool_calls", None)
                return reply, tool_calls

            except Exception as e:
                error = f"Error: {e}\n"
                self.Terminal(logger, error, "error")
                return error, None
            '''

        # Reply 
        response, function = get_response(model, memory, token, temperature, tools, True)  # response.strip()
        
        # Tools
        if function:
            output      = ""
            foutput     = ""
            message     = []
            drawn       = False

            for call in function:
                fname   = call.function.name
                args    = json.loads(call.function.arguments)

                # Add Media Features here
                if fname == "Weather":         output = self.Weather(args["location"])
                elif fname == "Music":         output = self.Music(args["song"])
                elif fname == "Youtube":       output = self.Youtube(args["video"])
                elif fname == "Spotify":       output = self.Spotify(args["song"])
                elif fname == "Note":          output = self.Note(args["reminder"], platform, user_id)
                elif fname == "Peek":          output = self.Peek(args["command"], platform, user_id)
                elif fname == "Currency":      output = self.Currency(args["from"], args["to"], args["amount"])
                elif fname == "Crypto":        output = self.Crypto(args["from"], args["to"], args["amount"])
                elif fname == "Time":          output = self.Time(args["from"], args["to"], args["time"])
                elif fname == "Shelly":        output = self.Shelly(args['device'], args['channel'], args['action'], args.get('duration')) # self.Shelly(**args) ?
                elif fname == "Web":           output = self.Web(name, user_id, chat_id, platform, args.get("query", ""), logger)   
                elif fname == "Maps":          output = self.Maps(**args)
                elif fname == "Audio":         output = self.Feature(name, user_id, chat_id, mediaDict[fname], fname,  args.get("state", "toggle"))
                elif fname == "Speak":         output = self.Feature(name, user_id, chat_id, mediaDict[fname], fname,  args.get("state", "toggle"))
                elif fname == "Vision":        output = self.Feature(name, user_id, chat_id, mediaDict[fname], fname,  args.get("state", "toggle"))
                elif fname == "Draw":          output = "" ; drawn  = True
                else:                          output = f"Unknown Tool: {fname}"
                if output not in (None, ""):   foutput += str(output) + "\n"   # if output: foutput += str(output) + "\n"
            
             # Only store tool text in memory if there's actually something
            if foutput.strip():     memory.append({self.Settings.gptRole: self.Settings.gptAssistant, self.Settings.gptName: name, self.Settings.gptContent: foutput.strip()})  # Is Memory Compound Necessary here? It's already Addded in prompt { foutput }
            if drawn and not (response or "").strip():  response = "🎨 Done."   # Send Drawn Boolean Back [ We send the tool back ] ( How about multi-tool call with Draw ? If Draw is first, it might not trigger. Send Tool List ?? If Draw is in multi-tool call, it won't trigger Synthesis ) { Draw breaks multi-tool call; Memory is appended with no output display message }
            if drawn == False:
                # MESSAGE NEEDS A SYSTEM MEMORY EDITION, HERE ~ 
                # Check by Platform
                # Memory mutates globally
                # Separate by Name to spawn multiple bots
                #  Is Product Overview relevant here ? Or Waste of Tokens ?? It's good to enforce the branding though ~ message.append({settings.gptRole: settings.gptSystem, settings.gptName: settings.FayName, settings.gptContent: settings.Product})
                #  if not synthesis or not str(synthesis).strip():   synthesis = "✓ Done."   
                #  if not foutput.strip():     response = "✓ Done."  { Tools ran but produced nothing textual – acknowledge, but you need to add error checker maybe ? }
                #  Printing Temperature Values for debugging, shrink to 2 Decimal Places
                
                # Copy
                prompt        = (f"""Rewrite the content. Dont say it is a rephrased\n
                                     Keep all URLs, Emojis, and important details. Use New Lines\n
                                     {foutput}""").strip()
                message       = memory[-3:]
                if platform == "Telegram":  message.append({self.Settings.gptRole: self.Settings.gptSystem, self.Settings.gptName: name, self.Settings.gptContent: self.Settings.Fay})
                elif platform == "Line":    message.append({self.Settings.gptRole: self.Settings.gptSystem, self.Settings.gptName: name, self.Settings.gptContent: self.Settings.FayLine})
                message.append({self.Settings.gptRole: self.Settings.gptAssistant, self.Settings.gptName: name, self.Settings.gptContent: prompt})
                
                # Synthesis    [ Claude might take long, causing connections to drop off, especially for Web]
                synthesis, _   = get_response(model, message, token, temperature, None, False)    # Should have a model argument to choose a smaller and cheaper model [ Might not need as you are using a small memory block to compute ]
                if synthesis:  response = synthesis ; memory.append({self.Settings.gptRole: self.Settings.gptAssistant, self.Settings.gptName: name, self.Settings.gptContent: response.strip()})  
        else:   memory.append({self.Settings.gptRole: self.Settings.gptAssistant, self.Settings.gptName: name, self.Settings.gptContent: response.strip()})
        return {
            "tool":      function,
            "response":  response,
            "memory":    memory         
        }

    def Draw(self, name, user_id, chat_id, platform, prompt, drawMemory, size, logger):
        """
        Channel-agnostic Draw helper.

        - Uses OpenAI DALL·E model from settings.Models["DallE"]["name"]
        - Appends to DrawMemoryDict[chat_id] to keep a running prompt context
        - Saves PNG under Media/<folder>/
        - Pushes the generated image into MemoryDict[chat_id] as a data URL
        - Returns a dict with image_path / image_url / timestamp or {"error": ...}
        """

        if not prompt or not str(prompt).strip():
            err = "Prompt is empty or null."
            if logger:  self.Terminal(logger, err, "error")
            return {"error": err}

        # Model + client (DALL·E is OpenAI-only, so we hardwire OpenAI here)
        model      = self.Settings.Models["DALLE"]["name"]
        client     = self.Settings.Clients["openai"]

        # Draw memory context
        drawMemory.append(prompt)
        image_prompt = "\n".join(drawMemory)

        # Optional log
        if logger:  self.Terminal(logger, f"{name} is Drawing Now ~ 📝")
        timestamp   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        try:
            # Generate image
            resp = client.images.generate(model=model, prompt=image_prompt, size=size, quality="hd")

            os.makedirs(os.path.join("Media", platform), exist_ok=True)

            image_url  = resp.data[0].url
            image_file = f"{name}_{chat_id}_Dalle.png"
            image_path = os.path.join("Media", platform, image_file)

            r = requests.get(image_url, timeout=30)
            r.raise_for_status()
            with open(image_path, "wb") as f:   f.write(r.content)
            #image       = media.FileEncoder(image_path)
            #memory.append({self.Settings.gptRole: self.Settings.gptAssistant, self.Settings.gptName: name, self.Settings.gptContent: [{"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{image}"}}]})
            
            ''' Memory URL is handled Outside
            self.MemoryDict[chat_id].append({
                self.Settings.gptRole: self.Settings.gptAssistant,
                self.Settings.gptName: name,
                self.Settings.gptContent: [
                    {"type": "text", "text": f"🎨 Sent image for prompt: {args['prompt']}"},
                    {"type": "image_url", "image_url": {"url": img_url}}
                ]
            })
            '''

            return {
                "image_path": image_path,
                "image_url":  image_url,
                "timestamp":  timestamp,
            }
        except Exception as e:
            err = f"Error (Draw): {e}"
            if logger:  self.Terminal(logger, err, "error")
            return {"error": err}

    def Web(self, name, user_id, chat_id, platform, prompt, logger):
        """
        Global Web tool for all adapters.
        Uses Responses API + web_search_preview.
        Returns plain text only (no citations in production).
        """
        if not prompt or not str(prompt).strip():   return "⚠️ Web: empty query."

        model      = self.Settings.Models["GPT5"]["name"]
        client     = self.Settings.Clients["openai"]

        try:
            resp = client.responses.create(model=model, tools=[{"type": "web_search"}], input=prompt)   # Perhaps we can confine to models gpt4.1 and up, we need an checker and a defaulter
            text = (getattr(resp, "output_text", "") or "").strip()
            if not text:    text = "⚠️ Web: no result text returned."

            ''' # Citations
            # Model + Client (prefer per-chat client if you keep that pattern)
            try:                model = self.Settings.Models["GPT41"]["name"]
            except Exception:   model = "gpt-4.1"

            client = None
            if chat_id is not None and hasattr(self, "ClientDict") and chat_id in self.ClientDict:  client = self.ClientDict[chat_id]
            else:   client = getattr(self, "Client", None) or getattr(self.Settings, "Clients", {}).get("openai")

            if client is None:  return "⚠️ Web: no OpenAI client available."
            
            # Debug / audit citations (kept here for dev use)
            citations = []
            for out in getattr(resp, "output", []) or []:
                for part in getattr(out, "content", []) or []:
                    for anno in getattr(part, "annotations", []) or []:
                        if getattr(anno, "type", "") == "url_citation":
                            url = getattr(anno, "url", None)
                            if url:
                                citations.append(url)

            citations = list(dict.fromkeys(citations))
            if citations:
                text += "\n\nCitations:\n" + "\n".join(citations[:8])
            '''

            return text
        except Exception as e:  return f"Error: {e}"

    def Feature(self, name, user_id, chat_id, mediaDict, feature, state):
        '''
        # pick the right dict
        if feature == "Audio":
            d = self.bAudioDict
            label = "Audio listening"
        elif feature == "Speak":
            d = self.bSpeakDict
            label = "Voice replies (TTS)"
        elif feature == "Vision":
            d = self.bImageDict
            label = "Vision"
        else:
            return f"Unknown feature: {feature}"

        # ensure key exists
        if chat_id not in d:    d[chat_id] = False

        # direct mutate (your preferred style)
        if state == "toggle":
            d[chat_id] = not d[chat_id]
        elif state == "on":
            d[chat_id] = True
        elif state == "off":
            d[chat_id] = False
        elif state == "status":
            pass
        return f"{label}: {'ON' if d[chat_id] else 'OFF'}"

        # Mutate
        if state == "toggle":   mediaDict = not mediaDict
        elif state == "on":     mediaDict = True
        elif state == "off":    mediaDict = False
        elif state == "status": pass
        '''

        # Mutation
        if state == "toggle":   mediaDict[chat_id] = not mediaDict[chat_id]
        elif state == "on":     mediaDict[chat_id] = True
        elif state == "off":    mediaDict[chat_id] = False
        elif state == "status": pass
        return f"{feature}: {'ON' if mediaDict[chat_id] else 'OFF'}"

    ########################################################################
    #########################         Media        #########################

    def Weather(self, city, *args):
        print("\n~~~~~~~~~  Weather Protocol Initiated  ~~~~~~~~~\n\n")
        city = ''.join(c for c in city if c not in string.punctuation) 

        output = ''

        url = f'http://api.openweathermap.org/data/2.5/weather?q={city}&appid={self.Settings.Weather_Key}'
        response = requests.get(url)
        if response.status_code != 200:
            output = "Oh dear, there seems to be an issue with the weather service, or the specified city may not exist."
            return output

        # Weather Data
        data = response.json()
        if 'weather' in data and 'main' in data:
            weather = str(data['weather'][0]['main'])
            weather = "Cloudy" if weather == "Clouds" else weather
            temperature = round(data['main']['temp'] - 273.15, 2)                   # Convert to Celsius
        else:
            output = "Weather data not available for the specified city."
            return output

        # Local Time
        localnow = datetime.now()

        # City Time
        now = datetime.utcnow()
        if 'timezone' in data:
            utc_now             = datetime.now(timezone.utc)
            city_timezone       = pytz.FixedOffset(data['timezone'] // 60)
            city_time           = utc_now.astimezone(city_timezone)
            formatted_city_time = city_time.strftime("%H:%M")
        else:
            output = "Error fetching city time."
            return output

        # Display
        sCity       = data['name']
        sDate       = str(now.month) + '/' + str(now.day) + '/' + str(now.year)
        sTime       = str(formatted_city_time)
        sWeather    = weather
        sTemp       = str(temperature) + '°C'
        output      = f"\t\t☀️\t\t🌦️\t\t⛈️\t\t❄️\t\t\n\n{sCity}\n{sDate}\n{sTime}\n{sWeather}\n{sTemp}"
        output      = f"\t\t☀️\t\t🌦️\t\t⛈️\t\t❄️\t\t\n\n{sWeather}\n{sTemp}\n{sCity}\n{sTime}\n{sDate}"
        output      = f"\t\t☀️\t\t🌦️\t\t⛈️\t\t❄️\t\t\n\n{sWeather}\n{sTemp}\n{sTime}\n{sDate}\n{sCity}"

        # [ - Experiment more with Chat Bubbles, and How they Work ~ ]
        '''        
        sCity = 'City: \t\t\t\t\t'       + data['name']
        sDate = 'Date:\t\t\t\t\t'        + str(now.day) + '/' + str(now.month) + '/' + str(now.year)
        sTime = 'Time:\t\t\t\t\t'        + str(formatted_city_time)
        sWeather = 'Weather: \t\t\t'     + weather
        sTemperature = 'Temperature: \t' + str(temperature) + '°C'
        '''

        print(output)
        return output

    def WeatherZ(self, city):
        # City
        city     = ''.join(c for c in city if c not in string.punctuation) 
        url      = f'http://api.openweathermap.org/data/2.5/weather?q={city}&appid={self.Settings.Weather_Key}'
        response = requests.get(url)
        if response.status_code != 200: return {"error": "Issue with the weather service or the specified city may not exist"}

        data = response.json()
        lat, lon = data['coord']['lat'], data['coord']['lon']

        # UV Index
        uv_url      = f'http://api.openweathermap.org/data/2.5/uvi?lat={lat}&lon={lon}&appid={self.Settings.Weather_Key}'
        uv_response = requests.get(uv_url)
        if uv_response.status_code != 200:  return {"error": "Issue with the UV Index service"}
        
        # Hourly
        hourly_url      = f'http://api.openweathermap.org/data/3.0/onecall?lat={lat}&lon={lon}&exclude=current,minutely,daily,alerts&appid={self.Settings.Weather_Key}'
        hourly_response = requests.get(hourly_url)
        if hourly_response.status_code != 200:  return {"error": "Issue with the hourly weather service"}
        
        hourly_data = hourly_response.json()

        hourly_weather = []
        for i, hour_data in enumerate(hourly_data['hourly'][:13]):
            hour    = datetime.fromtimestamp(hour_data['dt']).strftime('%H:%M')
            temp    = round(hour_data['temp'] - 273.15, 2)             # Convert Temperature to Celsius
            weather = str(hour_data['weather'][0]['main'])
            hourly_weather.append({"hour": hour, "temperature": f"{temp}°C", "weather": weather})

        # Weather
        if 'weather' in data and 'main' in data:
            weather     = str(data['weather'][0]['main'])
            weather     = "Cloudy" if weather == "Clouds" else weather
            humidity    = str(data['main']['humidity'])
            temperature = round(data['main']['temp'] - 273.15, 2)   # Convert Temperature to Celsius
            uv_index    = uv_response.json()['value']

            # UV
            if uv_index < 4:    uv = 'Low'
            elif uv_index < 8:  uv = 'Moderate'
            else:               uv = 'High'

            # Min-Max Temp
            if 'temp_min' in data['main'] and 'temp_max' in data['main']:
                temp_min = round(data['main']['temp_min'] - 273.15, 2)
                temp_max = round(data['main']['temp_max'] - 273.15, 2)
            else:
                temp_min, temp_max = "N/A", "N/A"

            # Wind
            if 'wind' in data and 'speed' in data['wind']:
                wind            = str(data['wind']['speed'])
                wind_deg        = str(data['wind']['deg'])
            else:
                wind, wind_deg  = "N/A", "N/A" 

            # Visibility
            if 'visibility' in data:    visibility = str(data['visibility'])
            else:                       visibility = "N/A"

            # Sunrise and Sunset Times
            sunrise = datetime.fromtimestamp(data['sys']['sunrise']).strftime('%H:%M')
            sunset = datetime.fromtimestamp(data['sys']['sunset']).strftime('%H:%M')
        else:   return {"error": "Weather data not available for the specified city"}

        # Output
        output = {
            "city": data['name'],
            "time": data['timezone'],
            "temperature": f"{temperature}°C",
            "Humidity": f"{humidity}",
            "UV": f"{uv}",
            "weather": weather,
            "sunrise": sunrise,
            "sunset": sunset,
            "wind": wind,
            "visibility": visibility,
            "mintemp": temp_min,
            "maxtemp": temp_max,
            "hourly_weather": hourly_weather}
        return output

    def Google(self, query, *args):
        print("\n~~~~~~~~~  Google Protocol Initiated  ~~~~~~~~~\n\n")
        baseURL = "https://www.google.com/search?q="
        search = query.replace(" ", "+")
        URL = baseURL + search
        return URL

    def Music(self, song, *args):
        print("\n~~~~~~~~~  Music Protocol  ~~~~~~~~~\n\n")
        ytmusic = YTMusic()
        search_results = ytmusic.search(song, filter="songs")
        try:
            song = search_results[0]
            video_id = song['videoId']
            URL = f'https://music.youtube.com/watch?v={video_id}'
        except Exception as e:  URL = f"\nWe have an Error: {e}"
        return URL

    def Youtube(self, video, *args):
        print("\n~~~~~~~~~ Youtube Protocol Initiated ~~~~~~~~~\n\n")
        try:
            results = YoutubeSearch(video, max_results=1).to_json()
            baseURL = "https://www.youtube.com/watch"
            suffixURL = re.search('/watch(.+?)"}]}', results).group(1)
            URL = baseURL + suffixURL
        except Exception as e:  URL = f"\nWe have an Error for Youtube: {e}"
        return URL

    def Spotify(self, song, *args):
        print("\n~~~~~~~~~ Spotify Protocol Initiated ~~~~~~~~~\n\n")
        try:
            search_results = self.Settings.SpotifyClient.search(q=song, type='track', limit=1)
            if search_results['tracks']['items']:
                track = search_results['tracks']['items'][0]
                track_id = track['id']
                URL = f'https://open.spotify.com/track/{track_id}'
        except Exception as e:  URL = f"\nWe have an Error: {e}"
        return URL

    def Wikipedia(self, query, *args):
        print("\n~~~~~~~~~  Wikipedia Protocol Initiated  ~~~~~~~~~\n\n")
        
        # Create an instance of the Wikipedia API
        wiki = wikipediaapi.Wikipedia(user_agent='Fay/1.0 (mostaste94@gmail.com)', language='en', extract_format=wikipediaapi.ExtractFormat.WIKI)

        # Search for the query
        page = wiki.page(query)

        # Check if the page exists
        if page.exists():
            output = page.fullurl
            output += page.summary[0:500] + '...'
        else:   output = "Sorry, I couldn't find any information on that topic."
        return output

    def News(self, category, *args):
        # Add More Sources !!!!!
        try:
            if not category == 'popular':
                #print(f"Fetching top headlines in {category} category.")
                top_headlines = self.Settings.NewsClient.get_top_headlines(language='en', category=category)
            else:
                #print("Fetching top headlines from Reuters and CNN.")
                top_headlines = self.Settings.NewsClient.get_top_headlines(language='en', sources='Reuters, CNN')

            top_headlines = pd.json_normalize(top_headlines['articles'])
            top_headlines.dropna(subset=['title','url', 'urlToImage', 'publishedAt'], inplace=True)

            topnewsdf = top_headlines[["title","url", "urlToImage", "publishedAt"]]
            dic = topnewsdf.set_index('title').T.to_dict('list')

            # Format the dictionary's keys and values in the desired structure ( Corresponding to {title, url, urlToImage, publishedAt} )
            output = [{"title": title, "content": values[0], "urlToImage": values[1], "publishedAt": values[2]} for title, values in dic.items()]
            return output
        except Exception as e:
            output = f"We have an Error: {e}"; print(output)
            return {"error": "Something went wrong with fetching the news!"}

    def Note(self, remind, folder, user_id, *args):
        print("\n~~~~~~~~~  Note Protocol Initiated  ~~~~~~~~~\n\n")
        reminder    = remind.strip()
        note        = os.path.join("Media", folder, f'{user_id}_Note.txt')
        with open(note, 'a', errors='ignore') as writer:
            writer.write(reminder + "\n")
            output = f"Reminder Noted: {reminder}"
        return output

    def Peek(self, command, folder, user_id, *args):
        print("\n~~~~~~~~~  Peek Protocol Initiated  ~~~~~~~~~\n\n")
        command = command.strip()
        note = os.path.join("Media", folder, f'{user_id}_Note.txt')
        if not os.path.exists(note):
            output = "Reminder does not exist.  Create Reminder First"
            return output

        if command == "read":
            with open(note, 'r', errors='ignore') as reader:
                data = reader.read()
                if data == '':  output = "No Reminders Found"
                else:           output = data
        elif command == "clear":
            with open(note, 'w', errors='ignore') as writer: pass
            output      = "Reminders Cleared"
        else:   output  = "Peek comes with 2 Commands: Read or Clear"
        return output

    def Time(self, from_city, to_city, specified_time):
        api_key = self.Settings.Time_Key
        base_url = "http://api.timezonedb.com/v2.1"
    
        def get_time_zone(city):
            url = f"{base_url}/get-time-zone?key={api_key}&format=json&by=zone&zone={city}"
            response = requests.get(url)
            return response.json().get('gmtOffset', 0)
        
        try:
            offset_from         = get_time_zone(from_city)
            offset_to           = get_time_zone(to_city)
            
            current_date        = datetime.now()
            specified_datetime  = datetime.strptime(specified_time, "%I:%M %p")
            combined_datetime   = current_date.replace(hour=specified_datetime.hour, minute=specified_datetime.minute, second=0, microsecond=0)
            
            time_difference     = (offset_to - offset_from)
            datetime_to         = combined_datetime + timedelta(seconds=time_difference)

            return f"{specified_time} in {from_city} is {datetime_to.strftime('%I:%M %p')} in {to_city}"
        
        except requests.exceptions.RequestException as e:   return str(e)

    def Crypto(self, from_crypto, to_currency, amount):
        print("\n~~~~~~~~~  Crypto Protocol Initiated  ~~~~~~~~~\n\n")
        url = f"https://api.coinbase.com/v2/prices/{from_crypto}-{to_currency}/spot"
        try:
            response    = requests.get(url)
            data        = response.json()
            rate        = float(data['data']['amount'])
            conversion  = rate * amount
            output      = f"{amount} {from_crypto} ≈ {conversion:.2f} {to_currency}\n Rate: {rate:.2f}"
            return output
        except requests.exceptions.RequestException as e:   return f"API Request Failed: {str(e)}"
        except KeyError:                                    return "Currency not supported"    

    def Currency(self, from_currency, to_currency, amount):
        print("\n~~~~~~~~~  Currency Protocol Initiated  ~~~~~~~~~\n\n")
        api_key = self.Settings.Currency_Key
        url     = f"https://api.exchangerate-api.com/v4/latest/{from_currency}"
        try:
            response    = requests.get(url)
            data        = response.json()
            rate        = data['rates'][to_currency]
            conversion  = rate * amount
            output      = f"{amount} {from_currency} ≈ {conversion:.2f} {to_currency}\n Rate: {rate:.2f}"
            return output
        except requests.exceptions.RequestException as e:   return f"API Request Failed: {str(e)}"
        except KeyError:                                    return "Currency not supported"

    def Expenses(self, date=None, source=None, description=None, debit=None, credit=None):
        excel = "Expenses.xlsx"
        
        wb = load_workbook(filename=excel)
        ws = wb.active
        
        # Find the next available row, starting at row 5
        next_row = 5
        while ws[f"A{next_row}"].value is not None:     next_row += 1

        print(f"Adding data on row {next_row}")  # Print the row number where data will be added

        # Append the new data in the respective columns
        ws[f"A{next_row}"] = date
        ws[f"B{next_row}"] = source
        ws[f"C{next_row}"] = description
        ws[f"D{next_row}"] = debit
        ws[f"E{next_row}"] = credit

        wb.save(excel)

        output = f"Expense added: {credit if credit != '0' else debit} for {description} on {date}"
        print(output)
        return output

    def ImageGen(self, name, memorydict, imageSize, platform, prompt=""):
        print("\n~~~~~~~~~  Draw Protocol Initiated  ~~~~~~~~~\n\n")
        MODEL           = self.Settings.Models["DallE"]['name']

        # These need Changes to wrap around 3 Different Platforms ( LINE, Tele, and Disc )
        # You have to Update Both Image and Normal Memory; In addition to that, you have to return the Image File for Sending.  So you gotta return 3 values
        # You can call the Parameter Memory, dont have to specify Dict, Pass in the specified Memory Object
        # Wait we dont need Platform, We just Return the Image and Update the Memory outside ?? Wait in this case we might not need Prompt either
        # BTW Image Generator from GPT 4.1 is Out, you can use Repsonses.Create

        if prompt:  memorydict.append(prompt)
        imagePrompt     = '\n'.join(memorydict)

        try:
            ##          Take out size? Maybe people dont really need it at this point
            #response   = self.Client.responses.create(model="gpt-4.1-mini", input=prompt, tools=[{"type": "image_generation"}])
            #response   = self.Client.images.generate(model=MODEL, prompt=imagePrompt, size=self.ImageSizeDict[chat_id], quality="hd")
            response    = self.Client.images.generate(model=MODEL, prompt=imagePrompt, size=imageSize, quality="hd")
            image_url   = response.data[0].url
            response    = requests.get(image_url)

            # This need changes ( Need to have a parameter to state which platform it is coming from )  
            # You also have to return multiple values ( This is Hard to Do.. The Whole Thing lmao )
            img_file    = os.path.join("Media", "Telegram", f'{self.Settings.InkName}Dalle_{name}.png')
            with open(img_file, 'wb') as out_file:     out_file.write(response.content)
            with open(img_file, 'rb') as photo_file:   return photo_file  # And updated Language and Image Memory 
        except Exception as e:  return f"Error: {e}\n"

    def Scrape(self, url):
        # Options
        options = FirefoxOptions()
        options.add_argument('--headless')
        options.page_load_strategy = "eager"            
        #   None / Normal / Eager
        ##  None - Normal loading, loads all javascript, videos, adware, you name it
        ##  Normal - Loads main DOM and minimal required JS related
        ##  Eager - Doesnt load JS related, loads only DOM and tags
        
        # Open Selenium
        driver = webdriver.Firefox(options=options)

        # Resolve short map links BEFORE selenium (more reliable than headless HTML)
        try:
            low = (url or "").lower()
            if "maps.app.goo.gl" in low or "goo.gl/maps" in low:    url = requests.get(url, allow_redirects=True, timeout=10, headers={"User-Agent": "Mozilla/5.0"}).url or url
        except Exception as e:  print(e)

        # Scrape Web
        try:
            driver.get(url)
            final_url = getattr(driver, "current_url", url) or url
            html      = driver.page_source
            u         = (final_url or "").lower()

            # ---------- MAPS ----------
            if any(k in u for k in ("google.com/maps", "maps.app.goo.gl", "goo.gl/maps", "maps.apple.com", "openstreetmap.org", "waze.com/ul", "wego.here.com", "here.com")):
                return Maps(driver).process_driver(final_url, html)

            # ---------- SPOTIFY ----------
            if "spotify" in u:
                song_name, artist_name = Spotify(driver).process_driver()
                return f"Song: {song_name}\nArtist: {artist_name}"

            # ---------- YOUTUBE MUSIC ----------
            if "music.youtube" in u:
                song_name, artist_name = YoutubeMusic(driver).process_driver()
                return f"Song: {song_name}\nArtist: {artist_name}"

            # ---------- ARTICLE / GENERAL ----------
            title, newsprompt = Article(driver).process_driver()
            return title, newsprompt
            
        except Exception as e:  print (e); return ""
        finally:                driver.quit()

    def Search(self, query, *args):
        print("\n~~~~~~~~~  SearchGPT Protocol Initiated  ~~~~~~~~~\n\n")
        baseURL = "https://www.google.com/search?q="
        search = query.replace(" ", "+")
        URL = baseURL + search
        return URL

    def Lyrics(self, song, artist=None):
        url = "https://genius.com/"
        options = FirefoxOptions()
        options.add_argument('--headless')
        options.page_load_strategy = 'none'         # We are going with None this time to load additional dom content due to clicks
        
        # Converts Url to UrlEncode
        song_query      = "+".join(song.split())
        artist_query    = "+".join(artist.split())
        search_query    = song_query + " " + artist_query
        url             = url + "search?q=" + search_query

        # WebDriver
        driver = webdriver.Firefox(options=options)
        driver.get(url)

        try:    wait = WebDriverWait(driver, timeout=5).until(lambda d : d.find_element(By.CLASS_NAME, "mini_card"))
        # Assuming the page does not load or the search result does not exist, Returns a no results screen or error screen
        except (SeleniumException.TimeoutException, SeleniumException.NoSuchElementException):
            page_source = driver.page_source
            soup = BeautifulSoup(page_source, "html.parser")
            text = '\n'.join(line for line in soup.get_text().split('\n') if line.strip())
            if "no results" in text.lower():    raise InterfaceException ("No results found")
            else:                               raise InterfaceException ("An error occurred (1)")
        
        # Clicks on the first item as requested
        wait.click()

        # Makes sure the lyrics and the top div is loaded before continuing
        try:    second_wait = WebDriverWait(driver, timeout=10).until(lambda d : d.find_element(By.CLASS_NAME, "iMpFIj") and d.find_element(By.ID, "lyrics-root"))
        
        # Failsafe in case the website doesn't load or whatever (marking it with a 2 to differentiate from the first error)
        except (SeleniumException.TimeoutException, SeleniumException.NoSuchElementException):
            raise InterfaceException ("An error occurred (2)")
        
        # Source
        page_source = driver.page_source

        # HTML
        soup = BeautifulSoup(page_source, "html.parser")

        # Song title class
        song_title = soup.find(class_="iMpFIj")

        # Artist title class
        artist_title = soup.find(class_="jhWHLb")

        if (song_title and artist_title):
            driver.close()

        song_title = song_title.get_text()
        artist_title = artist_title.get_text()
        body = ""
        body = "Song: " + song_title + "\n" + "Artist: " + artist_title + "\n\n"

        for p_tags in soup.find_all(class_="kUgSbL"):
            body += p_tags.get_text(separator='\n')
            body += "\n"
        return body

    def Stream(self, song_param, stream=False):
            """
            First to search for music (depending if they give direct url or just search context).
            Generally want to keep stream to false as packet drops occur quite often.

            Params:
            - `song_param` (string): The search query of the song

            Returns:
            The player object to be passed into discord 
            
            Usage:
            Pass to VLC Player or Discord Voice Channel
            """

            # Search 
            try:
                results     = YoutubeSearch(song_param, max_results=1).to_json()
                video_id    = re.search(r'"id":\s?"(.*?)"', results).group(1)  # Extract the video ID using regex
                URL         = f"https://www.youtube.com/watch?v={video_id}"
            except Exception as e:  URL = f"\nWe have an Error for Youtube: {e}"
            print (URL)

            player, filename = YTDLSource.from_url(URL, stream=stream)
            return player, filename

    def Shelly(self, device: str, channel: int, action: str, duration: int | None = None, debug: bool = False):
        action  = (action or "").lower()
        dev_key = (device or "").lower().strip()
        dev_key = self.ShellyAliases.get(dev_key, dev_key)

        if dev_key not in self.ShellyDevices:
            return {
                "ok": False,
                "error": f"Unknown Shelly device: {dev_key}",
                "allowed": list(self.ShellyDevices.keys())
            }

        host = self.ShellyDevices[dev_key]

        # Keep only the real hardware channel quirks
        override_channels = {
            "bedroom": 1,
            "balcony": 0,
        }
        if dev_key in override_channels:    channel = override_channels[dev_key]

        dev = self.ShellyManager.get_device(host)

        if not hasattr(self, "ShellyStateDict"):    self.ShellyStateDict = {}

        def _state_from(status: dict):
            is_on = status.get("output")
            if is_on is None:   is_on = status.get("on")
            return is_on

        def _get_status():  return dev.switch_get_status(channel)

        def _confirm(target: bool | None, tries: int = 4, delay: float = 0.15):
            last = None
            for _ in range(tries):
                st = _get_status()
                last = st
                cur = _state_from(st)
                if target is None or cur == target:  return cur, st, True
                time.sleep(delay)
            return _state_from(last) if last else None, last, False

        try:
            if action == "on":
                dev.switch_set(channel, True)
                is_on, status, confirmed = _confirm(True)
                self.ShellyStateDict[(host, channel)] = is_on
                out = {
                    "ok": confirmed,
                    "device": dev_key,
                    "host": host,
                    "channel": channel,
                    "state": "on" if is_on else "off",
                    "confirmed": confirmed
                }
                if debug:   out["status"] = status
                return out

            elif action == "off":
                dev.switch_set(channel, False)
                is_on, status, confirmed = _confirm(False)
                self.ShellyStateDict[(host, channel)] = is_on
                out = {
                    "ok": confirmed,
                    "device": dev_key,
                    "host": host,
                    "channel": channel,
                    "state": "on" if is_on else "off",
                    "confirmed": confirmed
                }
                if debug:   out["status"] = status
                return out

            elif action == "toggle":
                before = _state_from(_get_status())
                dev.switch_toggle(channel)
                target = None if before is None else (not before)
                is_on, status, confirmed = _confirm(target)
                self.ShellyStateDict[(host, channel)] = is_on
                out = {
                    "ok": confirmed,
                    "device": dev_key,
                    "host": host,
                    "channel": channel,
                    "state": "on" if is_on else "off",
                    "confirmed": confirmed
                }
                if debug:   out["status"] = status
                return out

            elif action == "status":
                status = _get_status()
                is_on = _state_from(status)
                self.ShellyStateDict[(host, channel)] = is_on
                out = {
                    "ok": True,
                    "device": dev_key,
                    "host": host,
                    "channel": channel,
                    "state": "on" if is_on else "off"
                }
                if debug:   out["status"] = status
                return out

            elif action == "power":
                status = _get_status()
                apower = status.get("apower")
                aenergy = (status.get("aenergy") or {}).get("total")
                out = {
                    "ok": True,
                    "device": dev_key,
                    "host": host,
                    "channel": channel,
                    "apower_w": apower,
                    "aenergy_wh_total": aenergy
                }
                if debug:   out["status"] = status
                return out

            elif action == "cover_open":
                dev.cover_open(0, duration=duration)
                return {
                    "ok": True,
                    "device": dev_key,
                    "host": host,
                    "action": "cover_open"
                }

            elif action == "cover_close":
                dev.cover_close(0, duration=duration)
                return {
                    "ok": True,
                    "device": dev_key,
                    "host": host,
                    "action": "cover_close"
                }

            elif action == "cover_stop":
                dev.cover_stop(0)
                return {
                    "ok": True,
                    "device": dev_key,
                    "host": host,
                    "action": "cover_stop"
                }

            else:
                return {
                    "ok": False,
                    "error": f"Unsupported Shelly action: {action}"
                }

        except ShellyError as e:
            return {
                "ok": False,
                "error": f"Shelly error: {e}",
                "device": dev_key,
                "host": host,
                "channel": channel
            }

    def Vector(self, vector, client, query: str, top_k: int = 3):
        """
        Semantic search over all loaded vector files (self.Media.Vectors).
        """

        if not (query or "").strip():   return "Usage: /search <query>"
        if not vector:                  return "No vectors loaded. Use /vector <name>."

        qvec = client.embeddings.create(model="text-embedding-3-small", input=query).data[0].embedding

        def cos(a, b):
            dot = sum(x*y for x, y in zip(a, b))
            na = math.sqrt(sum(x*x for x in a)) or 1e-12
            nb = math.sqrt(sum(y*y for y in b)) or 1e-12
            return dot / (na*nb)

        results = []

        for name, path in vector.items():
            try:
                with open(path, "r", encoding="utf-8") as f:    store = json.load(f)
                for rec in store.values():
                    if "embedding" in rec and "text" in rec:
                        score = cos(qvec, rec["embedding"])
                        results.append((score, name, rec["text"]))
            except Exception:   continue

        if not results: return "No embeddings found."

        results.sort(key=lambda x: x[0], reverse=True)

        out = []
        for score, src, txt in results[:top_k]:
            snippet = txt.replace("\n", " ").replace("\r", " ")
            out.append(f"{score:.4f} [{src}] :: {snippet}")
        return "\n\n".join(out)

    ########################################################################
    ########################          GPS           ########################

    # Maps
    def Maps(self, text=None, origin=None, destination=None, mode=None, near=None, waypoints=None):
        """
        Build a single Google Maps link. If we have origin/destination (or detect
        'from ... to ...' in text), return a DIRECTIONS URL; otherwise return SEARCH.

        # mode:         driving|walking|bicycling|transit
        # waypoints:    list[str]
        """
        
        # Normalize inputs
        t = (text or "").strip()
        waypoints = waypoints or []

        # ---- travel mode mapping -------------------------------------------------
        MODE_MAP = {
            "walk": "walking", "walking": "walking", "onfoot": "walking",
            "drive": "driving", "car": "driving", "motorcycle": "driving", "bike-taxi": "driving",
            "bike": "bicycling", "bicycle": "bicycling",
            "transit": "transit", "bus": "transit", "train": "transit", "metro": "transit",
            "bts": "transit", "mrt": "transit", "lrt": "transit", "skytrain": "transit"
        }

        # If no explicit mode, infer from text
        if not mode and t:
            m_mode = re.search(r"\b(walk(?:ing)?|drive|car|bike|bicycle|motorcycle|transit|bus|train|metro|bts|mrt|lrt|skytrain)\b", t, re.I)
            if m_mode:  mode = MODE_MAP.get(m_mode.group(1).lower(), "driving")
        mode = MODE_MAP.get(str(mode).lower(), "driving") if mode else "driving"

        # ---- detect directions intent -------------------------------------------
        ori, des = origin, destination

        # Pattern 1: "from A to B" (optionally "... via C, D")
        m = re.search(r"\bfrom\s+(.+?)\s+(?:to|->|→|＞|>|\u2192)\s+(.+?)(?:\s+\bvia\s+(.+))?$", t, re.I)
        if m:
            ori = ori or m.group(1).strip()
            des = des or m.group(2).strip()
            via = (m.group(3) or "").strip()
            if via:
                waypoints = waypoints or [] # split by common separators
                waypoints += [w.strip() for w in re.split(r"[|>,/;]+|(?:\s+and\s+)", via) if w.strip()]

        # Pattern 2: "directions to B" (origin left blank so Maps can use device location)
        if not des:
            m2 = re.search(r"\bdirections?\s+to\s+(.+)$", t, re.I)
            if m2:  des = m2.group(1).strip()

        # Pattern 3: arrow/short form "A → B" or "A - B"
        if not des:
            m3 = re.search(r"^\s*(.+?)\s*(?:->|→|＞|>|-|to)\s*(.+?)\s*$", t, re.I)
            if m3 and " near " not in t.lower():
                ori = ori or m3.group(1).strip()
                des = des or m3.group(2).strip()

        # If we clearly have a destination (and usually an origin), build DIRECTIONS
        if des:
            params = {
                "api": "1",
                "destination": des,
                "origin": ori or "",                # blank lets the app use current location
                "travelmode": mode,
            }
            if waypoints:   params["waypoints"] = "|".join(w for w in waypoints if w)
            return "https://www.google.com/maps/dir/?" + urllib.parse.urlencode(params, safe="|")

        # ---- otherwise: SEARCH intent -------------------------------------------
        # Try "near XYZ" tail
        if not near and t:
            m_near = re.search(r"\bnear\s+(.+)$", t, re.I)
            if m_near:  near = m_near.group(1).strip()

        query = t
        if near and near.lower() not in query.lower():  query = f"{(query + ' ').strip()}near {near}"
        return "https://www.google.com/maps/search/?api=1&query=" + urllib.parse.quote_plus(query or "Nearby")

    # Google    
    def GoogleMapsSmartLink(self, text: str):
        """
        Natural-language Google Maps linker:
        - "<kw> near <place>" or "<kw> nearby <place>" → search
        - "<from> -> <to> [driving|walking|bicycling|transit]" → directions
        - otherwise → place search
        Returns a user-ready Google Maps URL.
        """
        text = (text or "").strip()
        if not text:    return "Usage: /gmap <place | <kw> near <place> | <from> -> <to> [mode]>"

        q = requests.utils.quote

        low = text.lower()

        # --- directions: "<from> -> <to> [mode]" or "from A to B [mode]" ---
        if "->" in text:
            left, right = text.split("->", 1)
            origin = left.strip()
            toks   = right.strip().split()
            dest   = toks[0] if toks else right.strip()
            mode   = (toks[1] if len(toks) > 1 else "driving").lower()
            if mode not in ("driving","walking","bicycling","transit"):
                # treat token as part of destination if it's not a valid mode
                dest = " ".join(toks).strip()
                mode = "driving"
            return f"https://www.google.com/maps/dir/?api=1&origin={q(origin)}&destination={q(dest)}&travelmode={q(mode)}"

        if low.startswith("from ") and " to " in low:
            # "from A to B [mode]"
            body = text[5:]
            left, right = body.split(" to ", 1)
            origin = left.strip()
            toks   = right.strip().split()
            dest   = toks[0] if toks else right.strip()
            mode   = (toks[1] if len(toks) > 1 else "driving").lower()
            if mode not in ("driving","walking","bicycling","transit"):
                dest = " ".join(toks).strip()
                mode = "driving"
            return f"https://www.google.com/maps/dir/?api=1&origin={q(origin)}&destination={q(dest)}&travelmode={q(mode)}"

        # The LLM will only synthesize specified location and direction [ This function might be useful on Telegram ]
        '''
        # --- nearby: "<kw> near/nearby <place>" (Google accepts "keyword near place" as a query) ---
        for sep in (" near ", " nearby "):
            if sep in low:
                a, b = text.split(sep, 1)
                keyword = a.strip()
                place   = b.strip()
                # Google Maps doesn’t accept a radius parameter in the web URL; the plain query works great.
                query = f"{keyword} near {place}"
                return f"https://www.google.com/maps/search/?api=1&query={q(query)}"
        '''

        # --- default: place search ---
        return f"https://www.google.com/maps/search/?api=1&query={q(text)}"

    # Apple
    def AppleMapsSmartLink(self, text: str):
        """
        Natural-language Apple Maps linker:
        - "<kw> near <place>" or "<kw> nearby <place>" → search
        - "<from> -> <to> [driving|walking|bicycling|transit]" → directions
        - "from <A> to <B> [mode]" → directions
        - otherwise → place search
        Returns a user-ready Apple Maps URL.
        """
        text = (text or "").strip()
        if not text:
            return "Usage: /amap <place | <kw> near <place> | <from> -> <to> [mode]>"

        q = requests.utils.quote

        # Apple Maps dirflg: d (drive), w (walk), r (transit). :contentReference[oaicite:1]{index=1}
        # Bicycling isn't documented on that page, but is widely used as dirflg=b (availability varies by region). :contentReference[oaicite:2]{index=2}
        mode_map = {
            "driving":   "d",
            "walking":   "w",
            "transit":   "r",
            "bicycling": "b",
        }

        def _parse_dest_and_mode(s: str):
            """
            Accept:
            - "destination words... [mode]"
            If last token is a known mode, use it; otherwise default driving.
            """
            toks = (s or "").strip().split()
            if not toks:
                return "", "driving"

            last = toks[-1].lower()
            if last in mode_map:
                dest = " ".join(toks[:-1]).strip()
                mode = last
            else:
                dest = " ".join(toks).strip()
                mode = "driving"
            return dest, mode

        low = text.lower()

        # --- directions: "<from> -> <to> [mode]" ---
        if "->" in text:
            left, right = text.split("->", 1)
            origin = left.strip()
            dest, mode = _parse_dest_and_mode(right)

            dirflg = mode_map.get(mode, "d")
            return f"https://maps.apple.com/?saddr={q(origin)}&daddr={q(dest)}&dirflg={q(dirflg)}"

        # --- directions: "from A to B [mode]" (case-insensitive) ---
        m = re.match(r"^\s*from\s+(.+?)\s+to\s+(.+)\s*$", text, flags=re.IGNORECASE)
        if m:
            origin = m.group(1).strip()
            dest, mode = _parse_dest_and_mode(m.group(2))

            dirflg = mode_map.get(mode, "d")
            return f"https://maps.apple.com/?saddr={q(origin)}&daddr={q(dest)}&dirflg={q(dirflg)}"

        # --- search: "<kw> near <place>" or "<kw> nearby <place>" ---
        # Apple has a "near" param, but it’s meant as a coordinate hint in the archived docs,
        # so for human text we just keep it inside q (works nicely in practice). :contentReference[oaicite:3]{index=3}
        m2 = re.match(r"^\s*(.+?)\s+(near|nearby)\s+(.+)\s*$", text, flags=re.IGNORECASE)
        if m2:
            kw = m2.group(1).strip()
            place = m2.group(3).strip()
            return f"https://maps.apple.com/?q={q(f'{kw} near {place}')}"

        # --- default: place search ---
        return f"https://maps.apple.com/?q={q(text)}"

    ########################################################################
    ########################         Async          ########################

    async def WeatherAsync(self, message, city, *args):
        loop = asyncio.get_event_loop()
        output = await loop.run_in_executor(self.executor, self.Weather, city)
        await self.Reply(message, output)
        return output
    
    async def GoogleAsync(self, message, query, *args):
        loop = asyncio.get_event_loop()
        output = await loop.run_in_executor(self.executor, self.Google, query)
        await self.Reply(message, output)
        return output

    async def MusicAsync(self, message, song, *args):
        loop = asyncio.get_event_loop()
        output = await loop.run_in_executor(self.executor, self.Music, song)
        await self.Reply(message, output)
        return output

    async def YoutubeAsync(self, message, song, *args):
        loop = asyncio.get_event_loop()
        output = await loop.run_in_executor(self.executor, self.Youtube, song)
        await self.Reply(message, output)
        return output

    async def SpotifyAsync(self, message, song, *args):
        loop = asyncio.get_event_loop()
        output = await loop.run_in_executor(self.executor, self.Spotify, song)
        await self.Reply(message, output)
        return output

    async def WikipediaAsync(self, message, wiki, *args):
        loop = asyncio.get_event_loop()
        output = await loop.run_in_executor(self.executor, self.Wikipedia, wiki)
        await self.Reply(message, output)
        return output

    async def NoteAsync(self, message, remind, *args):
        loop = asyncio.get_event_loop()
        output = await loop.run_in_executor(self.executor, self.Note, remind, *args)
        await self.Reply(message, output)
        return output
    
    async def PeekAsync(self, message, command, *args):
        loop = asyncio.get_event_loop()
        output = await loop.run_in_executor(self.executor, self.Peek, command, *args)
        await self.Reply(message, output)
        return output

    async def SysAsync(self, message, *args):
        loop = asyncio.get_event_loop()
        output = await loop.run_in_executor(self.executor, self.Sys)
        await self.Reply(message, output)
        return output

    async def HelpAsync(self, message, *args):
        name = self.Settings.FayName
        loop = asyncio.get_event_loop()
        output = await loop.run_in_executor(self.executor, self.Help, name)
        await self.Reply(message, output)
        return output

    async def TimeAsync(self, message, from_zone, to_zone, time, *args):
        loop = asyncio.get_event_loop()
        output = await loop.run_in_executor(self.executor, self.Time, from_zone, to_zone, time)
        await self.Reply(message, output)
        return output

    async def CurrencyAsync(self, message, x, y, amount, *args):
        loop = asyncio.get_event_loop()
        output = await loop.run_in_executor(self.executor, self.Currency, x, y, amount)
        await self.Reply(message, output)
        return output
    
    async def CryptoAsync(self, message, x, y, amount, *args):
        loop = asyncio.get_event_loop()
        output = await loop.run_in_executor(self.executor, self.Crypto, x, y, amount)
        await self.Reply(message, output)
        return output

    async def ExpensesAsync(self, message, date, source, description, debit, credit, *args):
        loop = asyncio.get_event_loop()
        output = await loop.run_in_executor(self.executor, self.Expenses, date, source, description, debit, credit)
        await self.Reply(message, output)
        return output

    async def WebAsync(self, message, url, *args):
        loop = asyncio.get_event_loop()
        output = await loop.run_in_executor(self.executor, self.Web, url)
        await self.Reply(message, output)
        return output

    async def TerminalAsync(self, message, logger, text, type="info", *args):
        loop = asyncio.get_event_loop()
        await loop.run_in_executor(self.executor, self.Terminal, logger, text, type)
        return

    async def DrawAsync(self, message, name, chat_id, memorydict, prompt=""):
        loop = asyncio.get_event_loop()
        output = await loop.run_in_executor(self.executor, self.Draw, name, chat_id, memorydict, prompt="")
        await self.Reply(message, output)
        return

    async def LyricsAsync(self, message, song, artist, *args):
        loop = asyncio.get_event_loop()
        output = await loop.run_in_executor(self.executor, self.Lyrics, song, artist)
        await self.Reply(message, output)
        return
    
    async def ShellyAsync(self, message, device: str, channel: int, action: str, duration: int | None = None, debug: bool = False):
        loop = asyncio.get_event_loop()
        output = await loop.run_in_executor(self.executor, self.Shelly, device, channel, action, duration, debug)
        await self.Reply(message, output)
        return

    ########################################################################
    ########################          Coda          ########################

    # Utilities
    def _strip_markup(self, text: str) -> str:
        if not isinstance(text, str):   return text

        # **bold** and __bold__
        text = re.sub(r'\*\*(.+?)\*\*', r'\1', text, flags=re.DOTALL)
        text = re.sub(r'__(.+?)__',   r'\1', text, flags=re.DOTALL)

        # *italics* and _italics_
        # Guard against list bullets: require no whitespace immediately after the opening marker
        # and no whitespace immediately before the closing marker.
        text = re.sub(r'(?<!\w)\*(?!\s)(.+?)(?<!\s)\*(?!\w)', r'\1', text, flags=re.DOTALL)
        text = re.sub(r'(?<!\w)_(?!\s)(.+?)(?<!\s)_(?!\w)',   r'\1', text, flags=re.DOTALL)
        return text

    def _is_single_emoji(self, text: str) -> bool:    return bool(self.Emoji_Pattern.fullmatch((text or "").strip()))  # Or self.Emoji_Extended

    # Bearings
    def _looks_like_coords(self, s: str) -> bool:
        return bool(re.match(r"^\s*-?\d+(?:\.\d+)?\s*,\s*-?\d+(?:\.\d+)?\s*$", s or ""))

    def _normalize_lnglat(self, s: str) -> str | None:
        """Accept 'lng,lat' or 'lat,lng'; return 'lng,lat'."""
        try:                a, b = [float(x) for x in s.split(",")]
        except Exception:   return None
        # If the first looks like latitude (<=90) and second looks like longitude (<=180 but >90 often),
        # swap to make 'lng,lat'.
        lng, lat = a, b
        if abs(a) <= 90 and abs(b) <= 180 and abs(b) > 90:
            lng, lat = b, a
        if abs(lng) > 180 or abs(lat) > 90:
            return None
        return f"{lng},{lat}"

    def _ensure_lnglat(self, s: str) -> str | None:
        s = (s or "").strip()
        if self._looks_like_coords(s):  return self._normalize_lnglat(s)
        g = self.AMapGeocode(s)
        return g if (isinstance(g, str) and "," in g) else None

    def _normalize_tools_for_responses(self, tools_list):
        normalized = []
        for t in tools_list or []:
            # Old format: {"type":"function","function":{...}}
            if isinstance(t, dict) and t.get("type") == "function" and "function" in t:
                fn = t["function"] or {}
                normalized.append({
                    "type": "function",
                    "name": fn.get("name", "unnamed_tool"),
                    "description": fn.get("description", ""),
                    "parameters": fn.get("parameters", {"type": "object", "properties": {}}),
                })
            else:   normalized.append(t)    # Already in new format or something custom
        return normalized

    def MapsQuick(self, text, default_near=None):
        """
        Free-text -> ONE Google Maps URL.
        Examples:
        - 'Thai food near Ekkamai'      -> search
        - 'from Bang Na to ICONSIAM'    -> directions (driving by default)
        - 'BKK → Don Muang by train'    -> directions (mode=transit)
        """
        # Delegate to Maps, letting it auto-detect intent/mode.
        return self.Maps(text=text, near=default_near)

    # QQ Maps ( Tecent )
    def QQMapsSmartLink(self, text: str):
        """
        Natural-language QQ Maps linker (mirrors GoogleMapsSmartLink):
        - "<from> -> <to> [driving|walking|bicycling|transit]" → directions
        - "from A to B [mode]" → directions
        - otherwise → place search
        Returns a user-ready Tencent Maps URL (URI v1).
        """

        text = (text or "").strip()
        if not text:    return "Usage: /qqmap <place | <from> -> <to> [mode] | from A to B [mode]>"

        low = text.lower()
        key = "VSDBZ-BNMWM-5L66J-6DJHA-6QXDJ-ADFGL"
        referer = "Fay"

        def is_coords(s: str) -> bool:  return bool(re.match(r"^\s*-?\d+(?:\.\d+)?,\s*-?\d+(?:\.\d+)?\s*$", s or ""))

        mode_map = {
            "driving":   "drive",
            "walking":   "walk",
            "bicycling": "bike",
            "transit":   "bus",
        }

        # --- directions: "<from> -> <to> [mode]" ---
        if "->" in text:
            left, right = text.split("->", 1)
            origin = left.strip()
            toks   = right.strip().split()
            dest   = toks[0] if toks else right.strip()
            mode   = (toks[1] if len(toks) > 1 else "driving").lower()
            if mode not in mode_map:
                dest = " ".join(toks).strip()
                mode = "driving"

            base    = "https://apis.map.qq.com/uri/v1/routeplan"
            params  = {
                "type": mode_map[mode],
                "from": origin,
                "to": dest,
                "referer": referer,
                "key": key
            }
            if is_coords(origin):   params["fromcoord"] = origin.replace(" ", "")
            if is_coords(dest):     params["tocoord"] = dest.replace(" ", "")
            return f"{base}?{urlencode(params, safe=',')}"

        # --- directions: "from A to B [mode]" ---
        if low.startswith("from ") and " to " in low:
            body = text[5:]
            left, right = body.split(" to ", 1)
            origin = left.strip()
            toks   = right.strip().split()
            dest   = toks[0] if toks else right.strip()
            mode   = (toks[1] if len(toks) > 1 else "driving").lower()
            if mode not in mode_map:
                dest = " ".join(toks).strip()
                mode = "driving"

            base = "https://apis.map.qq.com/uri/v1/routeplan"
            params = {
                "type": mode_map[mode],
                "from": origin,
                "to": dest,
                "referer": referer,
                "key": key
            }
            if is_coords(origin):   params["fromcoord"] = origin.replace(" ", "")
            if is_coords(dest):     params["tocoord"] = dest.replace(" ", "")
            return f"{base}?{urlencode(params, safe=',')}"

        # --- default: place search ---
        base = "https://apis.map.qq.com/uri/v1/search"
        params = {
            "keyword": text,
            "referer": referer,
            "key": key
        }
        return f"{base}?{urlencode(params)}"

    # AMaps ( Gaode )
    def AMapSmartLink(self, text: str):
        text = (text or "").strip()
        if not text:    return "Usage: /amap <place | <kw> near <place> [radius] | <from> -> <to> [mode] | from A to B [mode]>"

        low = text.lower()

        # route: "<from> -> <to> [mode]"
        if "->" in text:
            left, right = text.split("->", 1)
            origin = left.strip()
            toks   = right.strip().split()
            dest   = toks[0] if toks else right.strip()
            mode   = (toks[1] if len(toks) > 1 else "driving").lower()
            return self.AMapRoute(origin, dest, mode)

        # route: "from A to B [mode]"
        if low.startswith("from ") and " to " in low:
            body = text[5:]
            left, right = body.split(" to ", 1)
            origin = left.strip()
            toks   = right.strip().split()
            dest   = toks[0] if toks else right.strip()
            mode   = (toks[1] if len(toks) > 1 else "driving").lower()
            return self.AMapRoute(origin, dest, mode)

        # nearby: "<kw> near/nearby <place> [radius|radiusm]"
        for sep in (" near ", " nearby "):
            if sep in low:
                a_low, b_low = low.split(sep, 1)
                keyword = text[:len(a_low)].strip()
                place   = text[len(a_low) + len(sep):].strip()

                parts = place.split()
                radius = None
                if parts:
                    last = parts[-1].lower()
                    if last.endswith("m") and last[:-1].isdigit():
                        radius = int(last[:-1])
                        place  = " ".join(parts[:-1]).strip()
                    elif last.isdigit():
                        radius = int(last)
                        place  = " ".join(parts[:-1]).strip()

                center = self._ensure_lnglat(place)
                if not center:  return f"https://uri.amap.com/search?keyword={quote(keyword)}"  # fallback to generic search instead of hard fail

                r = radius or 2000
                return f"https://uri.amap.com/search?keyword={quote(keyword)}&center={center}&radius={r}"

        # default: marker if we can geocode; else search
        loc = self._ensure_lnglat(text)
        if loc: return f"https://uri.amap.com/marker?position={loc}&name={quote(text)}"
        return f"https://uri.amap.com/search?keyword={quote(text)}"

    # Trip
    def TripLink(self,
                product="hotels",
                # --- Hotels ---
                city_id=None, city_name=None,
                checkin=None, checkout=None,           # "YYYY-MM-DD" or "YYYY/MM/DD"
                lat=None, lon=None,
                keyword=None,                          # e.g. "Hilton Sukhumvit Bangkok"
                hotel_id=None,                         # Trip's numeric hotel id if you have it
                rooms=1, adults=2, children=0,
                province_id=None, country_id=None, district_id=None,
                # --- Flights ---
                origin=None, destination=None,         # IATA (e.g. "SIN","BKK")
                depart_date=None,                      # "YYYY-MM-DD"
                triptype="OW",                         # OW | RT | MC
                cabin="Y",                             # Y|W|C|F
                quantity=1, childqty=0, babyqty=0,
                origin_name=None, destination_name=None,
                airline=None,                          # optional filter
                # --- Common ---
                locale=None, curr=None, ouid=None):
        """
        Build Trip.com deeplinks that mirror Trip's real surfaces,
        with affiliate params appended.
        """

        base    = "https://www.trip.com"
        product = (product or "hotels").strip().lower()

        # ---------- helpers ----------
        def _d(s):  # normalize dates to the formats Trip accepts in these endpoints
            if not s: return None
            s = str(s).replace("-", "/")
            try:
                # accept YYYY/MM/DD or YYYY-MM-DD; reformat to YYYY/MM/DD for hotels, keep YYYY-MM-DD for flights
                dt = datetime.strptime(s.replace("/", "-"), "%Y-%m-%d")
                return dt
            except Exception:
                return None

        def _fmt_h(dt):    return dt.strftime("%Y/%m/%d")
        def _fmt_f(dt):    return dt.strftime("%Y-%m-%d")

        def _add(d, **kv):
            for k,v in kv.items():
                if v is None or v == "" or v is False:  continue    # keep 0 if explicitly passed
                d[k] = v

        # Affiliate params (always)
        aff = {"Allianceid": "7007589", "SID": "260341649"}
        if ouid:   aff["ouid"] = str(ouid)
        if locale: aff["locale"] = locale
        if curr:   aff["curr"] = curr

        # ---------- HOTELS ----------
        if product in ("hotel","hotels"):
            # Dates: default to tomorrow +1 night if missing
            ci_dt = _d(checkin)
            co_dt = _d(checkout)
            today_th = datetime.now()  # (server tz OK — you can swap for Asia/Bangkok if needed)
            if not ci_dt:
                ci_dt = (today_th + timedelta(days=1))
            if not co_dt:
                co_dt = ci_dt + timedelta(days=1)

            params = {}
            # Core surfaces
            _add(params,
                city=city_id,
                cityName=city_name,
                provinceId=province_id or 0 if city_id else None,
                countryId=country_id,
                districtId=district_id or 0 if city_id else None,
                checkin=_fmt_h(ci_dt),
                checkout=_fmt_h(co_dt),
                lowPrice=0, highPrice=-1,                  # Trip uses these in examples
                crn=rooms, adult=adults, children=children,
                searchType="H",
                searchWord=keyword,
                barCurr=curr if curr else None)

            # Geo sprinkle (harmless if missing)
            if lat is not None and lon is not None:
                _add(params, lat=lat, lon=lon)
                # Coordinate bundle like their example
                google_coord = f"GOOGLE_{lat}_{lon}_0"
                normal_coord = f"NORMAL_{lat}_{lon}_0"
                _add(params, searchCoordinate=f"BAIDU_-1_-1_0|GAODE_-1_-1_0|{google_coord}|{normal_coord}")

            # If you know the hotel id, hint it like Trip does in SERP links
            if hotel_id:
                # pattern: searchValue = "31|{hotel_id}*31*{hotel_id}*1"
                sv = f"31|{hotel_id}*31*{hotel_id}*1"
                _add(params,
                    searchValue=sv,
                    searchBoxArg="t",
                    # Some SERP filters Trip adds; safe to omit—leave if you want Hilton-esque filtering:
                    # listFilters="29|1*29*1|2*2"
                )

            # Traffic source crumbs (optional; safe if omitted)
            _add(params, ctm_ref="ix_sb_dl", domestic=False)

            # Merge affiliate
            params.update(aff)

            return f"{base}/hotels/list?{urlencode(params, doseq=True)}"

        # ---------- FLIGHTS ----------
        if product in ("flight", "flights"):
            # Sanitize
            dcity = (origin or "").strip().lower() or None
            acity = (destination or "").strip().lower() or None

            # Default date: 7 days ahead if missing
            dd = _d(depart_date) or (datetime.now() + timedelta(days=7))

            params = {}
            _add(params,
                lowpricesource="lowPriceCalendar",
                triptype=triptype,
                _class=None,  # (not used; Trip expects 'class' below)
            )
            params["class"] = cabin  # 'class' is reserved in Python kwargs sometimes; set explicitly
            _add(params,
                quantity=int(quantity or 1),
                childqty=int(childqty or 0),
                babyqty=int(babyqty or 0),
                dcity=dcity, acity=acity,
                ddate=_fmt_f(dd),
                dcityName=origin_name,
                acityName=destination_name,
                airline=airline or "")

            # Merge affiliate bits last
            params.update(aff)

            return f"{base}/flights/showfarefirst?{urlencode(params, doseq=True)}"

        # ---------- ATTRACTIONS (fallback surface) ----------
        if product in ("attraction","attractions","things","tours"):
            params = {}
            _add(params, searchWord=keyword or city_name)
            params.update(aff)
            return f"{base}/things-to-do/?{urlencode(params, doseq=True)}"

        # Unknown → site root with affiliate tagging (still valid)
        return f"{base}/?{urlencode(aff, doseq=True)}"

    # CTrip
    def CTripLink(self,
                product="hotels",
                # --- Hotels ---
                city_id=None, city_name=None,
                checkin=None, checkout=None,           # "YYYY-MM-DD" or "YYYY/MM/DD"
                lat=None, lon=None,
                keyword=None,                          # e.g. "Hilton Sukhumvit Bangkok"
                hotel_id=None,                         # Ctrip/Trip numeric hotel id if you have it
                rooms=1, adults=2, children=0,
                province_id=None, country_id=None, district_id=None,
                # --- Flights ---
                origin=None, destination=None,         # IATA (e.g. "SIN","BKK")
                depart_date=None,                      # "YYYY-MM-DD"
                triptype="OW",                         # OW | RT | MC
                cabin="Y",                             # Y|W|C|F
                quantity=1, childqty=0, babyqty=0,
                origin_name=None, destination_name=None,
                airline=None,                          # optional filter
                # --- Common ---
                locale=None, curr=None, ouid=None):
        """
        Build Ctrip deeplinks (mirroring your Trip.com builder) with affiliate params appended.
        """
        # Ctrip bases per surface (you can collapse to a single base if you prefer)
        BASE_H = "https://hotels.ctrip.com"
        BASE_F = "https://flights.ctrip.com"
        BASE_R = "https://www.ctrip.com"

        # ---------- helpers ----------
        def _d(s):  # normalize dates to the formats the endpoints accept
            if not s: return None
            s = str(s).replace("-", "/")
            try:                dt = datetime.strptime(s.replace("/", "-"), "%Y-%m-%d");  return dt
            except Exception:   return None

        def _fmt_h(dt):    return dt.strftime("%Y/%m/%d")
        def _fmt_f(dt):    return dt.strftime("%Y-%m-%d")

        def _add(d, **kv):
            for k, v in kv.items():
                if v is None or v == "" or v is False:  continue    # keep 0 if explicitly passed
                d[k] = v

        # Affiliate params (always)
        aff = {"Allianceid": "7181658", "SID": "264218200"}
        if ouid:   aff["ouid"]   = str(ouid)
        if locale: aff["locale"] = locale
        if curr:   aff["curr"]   = curr

        p = (product or "hotels").strip().lower()

        # ---------- HOTELS ----------
        if p in ("hotel", "hotels"):
            # Dates: default to tomorrow +1 night if missing
            ci_dt = _d(checkin)
            co_dt = _d(checkout)
            today = datetime.now()
            if not ci_dt:   ci_dt = today + timedelta(days=1)
            if not co_dt:   co_dt = ci_dt + timedelta(days=1)

            params = {}
            _add(params,
                city=city_id,
                cityName=city_name,
                provinceId=province_id or (0 if city_id else None),
                countryId=country_id,
                districtId=district_id or (0 if city_id else None),
                checkin=_fmt_h(ci_dt),
                checkout=_fmt_h(co_dt),
                lowPrice=0, highPrice=-1,
                crn=rooms, adult=adults, children=children,
                searchType="H",
                searchWord=keyword,
                barCurr=curr if curr else None)

            # Optional lat/lon “sprinkle”
            if lat is not None and lon is not None:
                _add(params, lat=lat, lon=lon)
                google_coord = f"GOOGLE_{lat}_{lon}_0"
                normal_coord = f"NORMAL_{lat}_{lon}_0"
                _add(params, searchCoordinate=f"BAIDU_-1_-1_0|GAODE_-1_-1_0|{google_coord}|{normal_coord}")

            # If you know the hotel id, hint it like SERP links do
            if hotel_id:
                sv = f"31|{hotel_id}*31*{hotel_id}*1"
                _add(params, searchValue=sv, searchBoxArg="t")

            # Traffic source crumbs (optional)
            _add(params, ctm_ref="ix_sb_dl", domestic=False)

            # Merge affiliate
            params.update(aff)

            return f"{BASE_H}/hotels/list?{urlencode(params, doseq=True)}"

        # ---------- FLIGHTS ----------
        if p in ("flight", "flights"):
            dcity = (origin or "").strip().lower() or None
            acity = (destination or "").strip().lower() or None

            dd = _d(depart_date) or (datetime.now() + timedelta(days=7))

            params = {}
            _add(params, lowpricesource="lowPriceCalendar", triptype=triptype)
            params["class"] = cabin
            _add(params,
                quantity=int(quantity or 1),
                childqty=int(childqty or 0),
                babyqty=int(babyqty or 0),
                dcity=dcity, acity=acity,
                ddate=_fmt_f(dd),
                dcityName=origin_name,
                acityName=destination_name,
                airline=airline or "")

            params.update(aff)
            return f"{BASE_F}/flights/showfarefirst?{urlencode(params, doseq=True)}"

        # ---------- ATTRACTIONS (fallback surface) ----------
        if p in ("attraction", "attractions", "things", "tours"):
            params = {}
            _add(params, searchWord=keyword or city_name)
            params.update(aff)
            return f"{BASE_R}/things-to-do/?{urlencode(params, doseq=True)}"

        # Unknown → site root with affiliate tagging (still valid)
        return f"{BASE_R}/?{urlencode(aff, doseq=True)}"

    # Form [ Idea is great but not if they have bot blockers or json and captcha restrictions ] { Find a way to automate Captcha ~ }
    def WebOpenAndFill(self, url: str, profile: dict, field_map: list | None = None, headless: bool = False, submit_selector: dict | None = None, wait_sec: int = 20):
        
        # --- Open a page and fill fields from a profile JSON (Firefox) ---
        """
        Open url with Firefox and fill fields based on profile and field_map.
        - profile: your traveler profile dict (like the one you sent)
        - field_map: list of steps:
            {"by": "css"|"xpath", "selector": "...", "value": "{{contact.email}}"}
            {"click": true, "by": "css"|"xpath", "selector": "..."}   # for buttons/radios
        If omitted, this just opens the page (no fill).
        - submit_selector: optional click step to press Continue/Next at the end, e.g.:
            {"by":"xpath","selector":"//button[contains(.,'Continue') or @id='continue']"}
        - headless: run Firefox headless (False = visible for debugging)
        - wait_sec: explicit wait seconds for finding elements
        Returns a short status string.
        """

        # -------- small helpers --------
        def _mk_driver():
            opts = FirefoxOptions()
            if headless:    opts.headless = True
            # window size so headless layouts don’t break
            opts.add_argument("--width=1280")
            opts.add_argument("--height=900")
            service = FirefoxService(GeckoDriverManager().install())
            d = webdriver.Firefox(service=service, options=opts)
            d.set_page_load_timeout(60)
            d.implicitly_wait(2)
            return d

        def _by(kind: str):  return By.CSS_SELECTOR if (kind or "css").lower() == "css" else By.XPATH

        def _wait_find(drv, kind: str, sel: str):   return WebDriverWait(drv, wait_sec).until(EC.presence_of_element_located((_by(kind), sel)))

        def _wait_click(drv, kind: str, sel: str):
            el = WebDriverWait(drv, wait_sec).until(EC.element_to_be_clickable((_by(kind), sel)))
            el.click()
            return True

        # Resolve "{{a.b.c}}" from profile dict + derive convenience fields
        def _resolve(template: str) -> str:
            if not isinstance(template, str):
                return template
            if template.startswith("{{") and template.endswith("}}"):
                path = template[2:-2].strip()
                cur = _derived_profile
                for p in path.split("."):   cur = cur.get(p, "") if isinstance(cur, dict) else ""
                return "" if cur is None else str(cur)
            return template

        # Build a derived view (first/last name, dob parts, payment yyyy/mm, etc.)
        _derived_profile = dict(profile) if isinstance(profile, dict) else {}
        _derived_profile.setdefault("identity", {})
        _derived_profile.setdefault("passport", {})
        _derived_profile.setdefault("contact", {})
        _derived_profile.setdefault("payment", {})

        full = (_derived_profile["identity"].get("full_name") or "").strip()
        parts = [p for p in full.split() if p]
        if "first_name" not in _derived_profile["identity"]:
            _derived_profile["identity"]["first_name"] = parts[0] if parts else ""
        if "last_name" not in _derived_profile["identity"]:
            _derived_profile["identity"]["last_name"] = parts[-1] if len(parts) > 1 else ""

        dob = _derived_profile["identity"].get("dob", "")
        if len(dob) == 10 and dob.count("-") == 2:
            y, m, d = dob.split("-")
            _derived_profile["identity"].setdefault("dob_year", y)
            _derived_profile["identity"].setdefault("dob_month", m)
            _derived_profile["identity"].setdefault("dob_day", d)

        pay_exp = _derived_profile["payment"].get("expiry", "")
        if len(pay_exp) == 7 and pay_exp.count("-") == 1:
            yy, mm = pay_exp.split("-")
            _derived_profile["payment"].setdefault("expiry_year", yy)
            _derived_profile["payment"].setdefault("expiry_month", mm)

        driver = _mk_driver()

        '''
        driver = self._start_firefox_stealth(
                    headless=headless,
                    firefox_profile_path=None,   # or give your exact profile path if you want
                    user_agent=None,             # or set a real UA string
                    window_size=(1280, 900)
                )
        def _start_firefox_stealth(self, headless=False, firefox_profile_path=None, user_agent=None, window_size=(1280, 900)):
        """
        Launch Firefox with a real user profile (if available) and a few anti-bot prefs.
        - headless: False = show the browser so you can watch it work
        - firefox_profile_path: path to your real Firefox profile; if None, we try to auto-detect
        - user_agent: optional UA override (string)
        - window_size: (w, h)
        """
        import configparser

        def _detect_default_profile():
            # Try to find the default Firefox profile from profiles.ini
            if platform.system() == "Windows":
                base = os.path.join(os.environ.get("APPDATA", ""), "Mozilla", "Firefox")
            elif platform.system() == "Darwin":
                base = os.path.expanduser("~/Library/Application Support/Firefox")
            else:
                base = os.path.expanduser("~/.mozilla/firefox")
            ini = os.path.join(base, "profiles.ini")
            if not os.path.isfile(ini):
                return None
            cfg = configparser.ConfigParser()
            cfg.read(ini, encoding="utf-8")
            chosen = None
            for sec in cfg.sections():
                if cfg.has_option(sec, "Default") and cfg.get(sec, "Default") == "1":
                    chosen = cfg.get(sec, "Path", fallback=None)
                    is_rel = cfg.get(sec, "IsRelative", fallback="1") == "1"
                    if chosen:
                        return os.path.join(base, chosen) if is_rel else chosen
            # Fallback to first profile section
            for sec in cfg.sections():
                if cfg.has_option(sec, "Path"):
                    chosen = cfg.get(sec, "Path")
                    is_rel = cfg.get(sec, "IsRelative", fallback="1") == "1"
                    return os.path.join(base, chosen) if is_rel else chosen
            return None

        # Options
        opts = FirefoxOptions()
        if headless:
            opts.headless = True

        # Anti-bot-ish prefs
        # (Some are legacy, but harmless; helps on older builds)
        opts.set_preference("dom.webdriver.enabled", False)
        opts.set_preference("media.peerconnection.enabled", False)           # reduce WebRTC leaks
        opts.set_preference("privacy.trackingprotection.enabled", True)
        opts.set_preference("network.http.referer.XOriginPolicy", 1)

        if user_agent:
            opts.set_preference("general.useragent.override", user_agent)

        # Window size (helps headless layouts)
        w, h = window_size
        opts.add_argument(f"--width={int(w)}")
        opts.add_argument(f"--height={int(h)}")

        # Attach a real Firefox profile if available (cookies/extensions/fonts look more human)
        ff_profile = None
        try:
            from selenium.webdriver.firefox.firefox_profile import FirefoxProfile
            if not firefox_profile_path:
                firefox_profile_path = _detect_default_profile()
            if firefox_profile_path and os.path.isdir(firefox_profile_path):
                ff_profile = FirefoxProfile(firefox_profile_path)
                ff_profile.set_preference("dom.webdriver.enabled", False)
                if user_agent:
                    ff_profile.set_preference("general.useragent.override", user_agent)
        except Exception:
            ff_profile = None  # non-fatal

        service = FirefoxService(GeckoDriverManager().install())
        driver = webdriver.Firefox(service=service, options=opts, firefox_profile=ff_profile)   # Firefox_Profile is an unexpected Keyword Argument
        driver.set_page_load_timeout(60)
        driver.implicitly_wait(2)

        # Light-touch JS patch: unset navigator.webdriver (works on some builds)
        try:
            driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined});")
        except Exception:
            pass

        return driver
        '''

        try:
            driver.get(url)

            if not field_map:   return "Opened page (no field_map provided)."

            for step in field_map:
                if step.get("click"):   _wait_click(driver, step.get("by", "css"), step["selector"]); continue

                # fill
                kind = step.get("by", "css")
                sel  = step["selector"]
                raw  = step.get("value", "")
                val  = _resolve(raw)

                el   = _wait_find(driver, kind, sel)
                
                # attempt to clear safely
                try:                el.clear()
                except Exception:   pass
                el.send_keys(val)
                # blur to trigger onChange
                try:                el.send_keys(Keys.TAB)
                except Exception:   pass

            if submit_selector: _wait_click(driver, submit_selector.get("by", "css"), submit_selector["selector"])
            return "✅ WebOpenAndFill completed."
        except Exception as e:  return f"Error during fill: {e}"
        finally:
            try:
                if headless:    driver.quit()   # keep the browser open for manual review if not headless
            except Exception:   pass

    # Misc
    def AMapGeocode(self, address: str):
        key = "57aeece95915b8fa9a7e7d8c5d524a8b"
        if not key: return "Amap key missing. Set Settings.AMap_Key or AMAP_KEY."
        try:
            r = requests.get(
                "https://restapi.amap.com/v3/geocode/geo",
                params={"address": address, "key": key, "output": "JSON"},
                timeout=8,
            )
            data = r.json()
            if data.get("status") != "1" or not data.get("geocodes"):   return f"AMapGeocode error: {data.get('info','') or 'no result'}"
            loc = data["geocodes"][0].get("location")  # 'lng,lat'
            norm = self._normalize_lnglat(loc or "")
            return norm or loc
        except Exception as e:  return f"AMapGeocode error: {e}"

    def AMapRoute(self, origin: str, destination: str, mode: str = "driving", city: str | None = None):
        key = "57aeece95915b8fa9a7e7d8c5d524a8b"
        if not key: return "Amap key missing. Set Settings.AMap_Key or AMAP_KEY."

        o = self._ensure_lnglat(origin)
        d = self._ensure_lnglat(destination)
        if not (o and d): return "Could not resolve origin/destination."

        mode = (mode or "driving").lower()
        mode_map = {"driving":"car","walking":"walk","bicycling":"ride","transit":"bus"}
        link = f"https://uri.amap.com/navigation?from={o}&to={d}&mode={mode_map.get(mode,'car')}"

        if mode == "transit" and not city:
            return f"Transit • (open in Amap)\n{link}"

        try:
            if mode == "driving":
                url = "https://restapi.amap.com/v3/direction/driving"
                params = {"origin": o, "destination": d, "key": key}
            elif mode == "walking":
                url = "https://restapi.amap.com/v3/direction/walking"
                params = {"origin": o, "destination": d, "key": key}
            elif mode == "bicycling":
                url = "https://restapi.amap.com/v4/direction/bicycling"
                params = {"origin": o, "destination": d, "key": key}
            else:
                url = "https://restapi.amap.com/v3/direction/transit/integrated"
                params = {"origin": o, "destination": d, "city": city, "key": key}

            data = requests.get(url, params=params, timeout=10).json()
            if data.get("status") != "1":   return f"AMapRoute error: {data.get('info','') or 'no result'}"

            distance_m, duration_s = 0, 0
            if mode in ("driving","walking"):
                paths = (data.get("route") or {}).get("paths") or []
                if not paths: return "AMapRoute error: no route"
                first = paths[0]
                distance_m = int(float(first.get("distance", 0)))
                duration_s = int(float(first.get("duration", 0)))
            elif mode == "bicycling":
                paths = (data.get("data") or {}).get("paths") or []
                if not paths: return "AMapRoute error: no route"
                first = paths[0]
                distance_m = int(float(first.get("distance", 0)))
                duration_s = int(float(first.get("duration", 0)))
            else:
                transits = (data.get("route") or {}).get("transits") or []
                if not transits: return "AMapRoute error: no route"
                t0 = transits[0]
                distance_m = int(float(t0.get("distance", "0")))
                duration_s = int(float(t0.get("duration", "0")))

            km = round(distance_m/1000, 2)
            mins = max(1, duration_s // 60)
            return f"{mode.title()} • ~{km} km • ~{mins} min\n{link}"
        except Exception as e:  return f"AMapRoute error: {e}"

    def AMapNearby(self, keyword: str, location_lnglat: str, radius: int = 2000):
        key = "57aeece95915b8fa9a7e7d8c5d524a8b"
        if not key: return "Amap key missing. Set Settings.AMap_Key or AMAP_KEY."
        center = self._normalize_lnglat(location_lnglat)
        if not center: return f"Invalid center: {location_lnglat}"

        try:
            r = requests.get(
                "https://restapi.amap.com/v5/place/around",
                params={
                    "key": key,
                    "types": "",
                    "keywords": keyword,
                    "location": center,   # 'lng,lat'
                    "radius": int(radius),
                    "page_size": 10,
                    "page_num": 1,
                },
                timeout=10,
            )
            data = r.json()
            if data.get("status") != "1":   return f"AMapNearby error: {data.get('info','') or 'no result'}"

            pois = (data.get("pois") or [])[:5]
            if not pois: return "No places found."

            lines = []
            for i, p in enumerate(pois, 1):
                name = (p.get("name") or "").strip()
                addr = (p.get("address") or "").strip()
                loc  = self._normalize_lnglat(p.get("location") or "")
                link = f"https://uri.amap.com/marker?position={loc}&name={quote(name)}" if loc else ""
                lines.append(f"{i}. {name}\n   {addr}\n   {loc or '(no coords)'}\n   {link}")
            return "\n".join(lines)
        except Exception as e:  return f"AMapNearby error: {e}"

    def TripAff(self, url, ouid=None, locale="en-XX", curr="USD"):
        """
        Take ANY Trip.com URL (path or full) and append/overwrite affiliate params.
        Returns a single final URL string.
        """
        print("\n~~~~~~~~~  Trip Affiliate Protocol Initiated  ~~~~~~~~~\n")
        try:
            # If they pass a path like "/hotels/..." normalize to full URL
            if not re.match(r"^https?://", str(url), flags=re.I):
                path = ("/" + str(url).lstrip("/")) if url else "/"
                url = f"https://www.trip.com{path}"

            parsed = urlparse(url)  # Only touch trip.com links; otherwise return untouched
            if "trip.com" not in parsed.netloc.lower(): return url

            # Merge existing query with our affiliate params
            q = dict(parse_qsl(parsed.query, keep_blank_values=True))
            q["Allianceid"] = "7007589"
            q["SID"]        = "260341649"
            if ouid:   q["ouid"]   = str(ouid)
            if locale: q["locale"] = locale
            if curr:   q["curr"]   = curr

            new_query = urlencode(q, doseq=True)
            final     = urlunparse((parsed.scheme or "https",
                                    parsed.netloc,
                                    parsed.path or "/",
                                    parsed.params,
                                    new_query,
                                    parsed.fragment))
            return final
        except Exception as e:  return f"Error (TripAff): {e}"

        # --- CTrip ( Mini-Apps ) { We need their API } ---
        # Web Scrapper to build Database manually [ Localization ]

        # --- Google Maps dynamic links (node.py) ---
        # One link out, always. Builds either a SEARCH or a DIRECTIONS URL (api=1).
        # No external deps; robust regex parsing for "from X to Y", "near Z", mode, waypoints.
        # Maps Link is Different for China

    # Decode Web Link to Geo Coordinate for Maps related Share [ Relative to Platforms, Share from Browsers and App and OS is Different ] { DONE }
    ''' Classic Shelly
    def Shelly(self, device: str, channel: int, action: str, duration: int | None = None, debug: bool = False):
        # Config
        def ShellyIP(name: str):
            name = name.lower().replace(" ", "_")
            aliases = {
                "library_main": "library",
                "library_lights": "library",
                "bedroom_light": "bedroom",
                "bedroom_lights": "bedroom",
                "bathroom_b": "bathroom",
                "bathroom_light": "bathroom",
                "bathroom_lights": "bathroom",
                "bathroom_l": "restroom",
                "restroom": "restroom",
                "restroom_lights": "restroom",
                "lavatory": "restroom",
                "toilet": "restroom",
                "closet_lights": "closet",
                "balcony_lights": "balcony",
                "sky_lights": "sky",
                "deck_lights": "deck",
                "loft_lights": "loft",
                "sky": "sky",
                "deck": "deck",
                "loft": "loft",
                "caelum": "loft",
                "libram": "library",
            }


            # Latin
            aliases = {

                # --- Faux-Latin / spell-ish aliases (device targets) ---
                # (pick whatever vocabulary you actually use)
                "cubiculum": "bedroom_main",      # bedroom
                "bibliotheca": "library_main",    # library
                "latrina": "bedroom_bathroom",    # bathroom (bedroom side)
                "latrina_libris": "library_bathroom",
                "armarium": "bedroom_closet",     # closet
                "balconeum": "balcony",           # balcony
                "caelum": "sky",                  # sky
                "solarium": "deck",               # deck / terrace
                "cenaculum": "loft",              # loft
            }

            # Properties
            for suffix in ("_lights", "_light", "_lamp", "_switch"):
                if name.endswith(suffix):
                    base = name[:-len(suffix)]
                    if base in self.ShellyDevices:
                        name = base
                        break

            if name in aliases: name = aliases[name]
            if name not in self.ShellyDevices: return None
            return self.ShellyDevices[name]

        dev_key = (device or "").lower().replace(" ", "_")
        host    = ShellyIP(dev_key) or device

        override_channels = {
            "bedroom_master": 1,
            "bedroom_main": 1,
            "bedroom": 1,
            "balcony": 0,
            "balcony_light": 0,
            "bedroom_balcony": 0,
        }
        if dev_key in override_channels:    channel = override_channels[dev_key]

        dev    = self.ShellyManager.get_device(host)
        action = (action or "").lower()

        if not hasattr(self, "ShellyStateDict"):    self.ShellyStateDict = {}

        def _state_from(status: dict):
            is_on = status.get("output")
            if is_on is None: is_on = status.get("on")
            return is_on

        def _get_status():
            return dev.switch_get_status(channel)

        def _confirm(target: bool | None, tries: int = 4, delay: float = 0.15):
            last = None
            for _ in range(tries):
                st = _get_status()
                last = st
                cur = _state_from(st)
                if target is None or cur == target:
                    return cur, st, True
                time.sleep(delay)
            return _state_from(last) if last else None, last, False

        try:
            # Lights
            if action == "on":
                dev.switch_set(channel, True)
                is_on, status, confirmed = _confirm(True)
                self.ShellyStateDict[(host, channel)] = is_on
                out = {"ok": confirmed, "device": dev_key, "host": host, "channel": channel, "state": "on" if is_on else "off", "confirmed": confirmed}
                if debug: out["status"] = status
                return out

            elif action == "off":
                dev.switch_set(channel, False)
                is_on, status, confirmed = _confirm(False)
                self.ShellyStateDict[(host, channel)] = is_on
                out = {"ok": confirmed, "device": dev_key, "host": host, "channel": channel, "state": "on" if is_on else "off", "confirmed": confirmed}
                if debug: out["status"] = status
                return out

            elif action == "toggle":
                # optional: read before to know the target
                before = _state_from(_get_status())
                dev.switch_toggle(channel)
                target = None if before is None else (not before)
                is_on, status, confirmed = _confirm(target)
                self.ShellyStateDict[(host, channel)] = is_on
                out = {"ok": confirmed, "device": dev_key, "host": host, "channel": channel, "state": "on" if is_on else "off", "confirmed": confirmed}
                if debug: out["status"] = status
                return out

            elif action == "status":
                status = _get_status()
                is_on  = _state_from(status)
                self.ShellyStateDict[(host, channel)] = is_on
                out = {"ok": True, "device": dev_key, "host": host, "channel": channel, "state": "on" if is_on else "off"}
                if debug: out["status"] = status
                return out

            elif action == "power":
                status  = _get_status()
                apower  = status.get("apower")
                aenergy = (status.get("aenergy") or {}).get("total")
                out = {"ok": True, "device": dev_key, "host": host, "channel": channel, "apower_w": apower, "aenergy_wh_total": aenergy}
                if debug: out["status"] = status
                return out

            # Roller
            elif action == "cover_open":
                dev.cover_open(0, duration=duration)
                return {"ok": True, "device": dev_key, "host": host, "action": "cover_open"}

            elif action == "cover_close":
                dev.cover_close(0, duration=duration)
                return {"ok": True, "device": dev_key, "host": host, "action": "cover_close"}

            elif action == "cover_stop":
                dev.cover_stop(0)
                return {"ok": True, "device": dev_key, "host": host, "action": "cover_stop"}

            else:                   return {"ok": False, "error": f"Unsupported Shelly action: {action}"}
        except ShellyError as e:    return {"ok": False, "error": f"Shelly error: {e}", "device": dev_key, "host": host, "channel": channel}
    '''

########################################################################
########################        I   /   O       ########################

# Discord
class Engine(Interface):
    def __init__(self, settings):   super().__init__(settings)
        
    async def Reply(self, message, response):
        await self.Message(message.channel, response)
        pass

    async def Message(self, channel, text, chunk_size=2000):
        if len(text) <= chunk_size: await channel.send(text)
        else:
            start = 0
            end = chunk_size
            while start < len(text):
                await channel.send(text[start:end])
                start += chunk_size
                end += chunk_size

    async def Save(self, chat_id, user_id, user_name, message_id, message, timestamp, content="Text", file_name=None, file_path=None):
        if content == "Text":
            message_document = {
                'user_id':      user_id,
                'user_name':    user_name,
                'message_id':   message_id,
                'message':      message,
                'content':      content,
                'timestamp':    timestamp
            }
        else:
            #with open(file_path, "rb") as file_data:    file_id = self.Settings.MongoFS.put(file_data, filename=file_name)    # Store the media in GridFS
            message_document = {
                'user_id':      user_id,
                'chat_id':      chat_id,
                'message_id':   message_id,
                'file_path':    file_path,               # Change to file_id for Mongo
                'file_name':    file_name,
                'content':      content,
                'timestamp':    timestamp
            }

        # ⬆️ User save
        user_db = self.Settings.DiscordMessages.user_db(chat_id)       # Per-user ledger
        user_db.insert_one(message_document)

        # ⬇️ Swap out Mongo with local JSONL
        self.Settings.DiscordMessages.insert_one("DiscordMessages", message_document)
        print(f"Message added to chat {chat_id} (Offline DB)\n")

        # MongoDB
        ##result = self.Settings.DiscordMessages.update_one({"chat_id": chat_id}, {"$push": {"messages": message_document}}, upsert=True); print(f"Message added to chat {chat_id} with result: {result.modified_count}\n")
        return

    async def MusicStream(self, song_param, stream=False):
        """
        First to search for music (depending if they give direct url or just search context).
        Generally want to keep stream to false as packet drops occur quite often.

        Params:
        - `song_param` (string): The search query of the song

        Returns:
        The player object to be passed into discord engine
        """

        # Search 
        try:
            results     = YoutubeSearch(song_param, max_results=1).to_json()
            video_id    = re.search(r'"id":\s?"(.*?)"', results).group(1)  # Extract the video ID using regex
            URL         = f"https://www.youtube.com/watch?v={video_id}"
        except Exception as e:  URL = f"\nWe have an Error for Youtube: {e}"
        print (URL)

        player, filename = await YTDLSource.from_url(URL, stream=stream)
        return player, filename

# Telegram 
class Node(Interface):
    def __init__(self, settings, blueprint):
        super().__init__(settings)
        self.blueprint = blueprint

    def Reply(self, update, context, response, markup=None):
        # Response Check
        response = (response or "").strip()
        if not response:    response = "✓ Done."

        # Reply Counter
        counter = random.uniform(0.0, 1.0)
        chat_id = update.effective_chat.id

        # Update Message
        if hasattr(update, 'message'):
            # Counter to tag back replies 
            if counter < 0.35:  update.message.reply_text(response, reply_markup=markup)
            else:               context.bot.send_message(chat_id=chat_id, text=response, reply_markup=markup)
        else:                   context.bot.send_message(chat_id=chat_id, text=response, reply_markup=markup)
        pass

        # Edit Messages 
        '''
        query       = update.callback_query
        query.edit_message_text(text=response, reply_markup=markup)
        message     = context.bot.send_message(chat_id=update.effective_chat.id, text=text)
        message_id  = message.message_id
        #context.bot.edit_message_text(chat_id=chat_id, message_id=message_id, text="New text", reply_markup=new_markup)
        '''

    def Message(self, update, context, response, chunk_size=2000):
        # Vincent work on this soon
        # Added a fixed number 50 for the bracket indicating which paragraph
        # Is this even relevant ? it will reply within 4096.  But I saw that a human user paste more than that and it sent in chunks.  Not sure if possible for Bot
        # Just cut the text in 2, then send twice ? 
        # Wait wth is this function for ? Message is needed in Discord because every message has a Chunk; did we even use this function at all ? This Function sort of become oboselete
        # No one wants to Read that much on Chat Platforms
        true_chunk_size = chunk_size - 20
        num_of_chunks = len(response) // true_chunk_size

        if num_of_chunks <= 1:
            context.bot.send_message(chat_id=update.effective_chat.id, text=response)
            return
        
        if len(response) % chunk_size:  num_of_chunks += 1

        chunks = [response[i * true_chunk_size : (i+1) * true_chunk_size] for i in range(num_of_chunks)]
        
        for index, chunk in enumerate(chunks):
            chunk += f"\n({index+1} out of {num_of_chunks})"
            context.bot.send_message(chat_id=update.effective_chat.id, text=chunk)
            time.sleep(1)

    def Save(self, chat_id, user_id, user_name, message_id, message, timestamp, content="Text", file_name=None, file_path=None):
        if content == "Text":
            message_document = {
                'user_id':      user_id,
                'user_name':    user_name,
                'message_id':   message_id,
                'message':      message,
                'content':      content,
                'timestamp':    timestamp
            }
        else:
            #with open(file_path, "rb") as file_data:    file_id = self.Settings.MongoFS.put(file_data, filename=file_name)    # Store the media in GridFS
            message_document = {
                'user_id':      user_id,
                'user_name':    user_name,
                'message_id':   message_id,
                'chat_id':      chat_id,
                'file_path':    file_path,               # Change to file_id for Mongo ( Should you also include the b64 ? )
                'file_name':    file_name,
                'content':      content,
                'timestamp':    timestamp
            }
        
        # ⬆️ User save
        user_db = self.Settings.TelegramMessages.user_db(chat_id)       # Per-user ledger
        user_db.insert_one(message_document)

        # ⬇️ Swap out Mongo with local JSONL
        self.Settings.TelegramMessages.insert_one("TeleMessages", message_document)
        print(f"Message added to chat {chat_id} (Offline DB)\n")

        # MongoDB
        ##result = self.Settings.TeleMessages.update_one({"chat_id": chat_id}, {"$push": {"messages": message_document}}, upsert=True); print(f"Message added to chat {chat_id} with result: {result.modified_count}\n")
        return

    def Listener(self, *args):
        """
        Run in background. Whenever you type a single `m` + Enter, it will call multicast_prompt().
        """
        while True:
            line = sys.stdin.readline().strip()
            if line.lower() == "m":
                print("\n🔧 [stdin] 'm' detected → launching Multicast prompt…\n")
                self.blueprint.Multicast(bTerminal=True)
            # else: ignore other input

    def Weather(self, city):
        output = super().Weather(city)
        return output
    
# Line
class Hub(Interface):
    def __init__(self, settings, greenprint):   
        super().__init__(settings)
        self.greenprint = greenprint
    
    def Reply(self, event, line, response, image=False, quick=None, emojis=None):
        # Perhaps you should make this Function be able to accept all Message Types ? If you send in the ready made Types, you dont need Image parameter
        if image:   line.reply_message(event.reply_token, ImageSendMessage(original_content_url=response, preview_image_url=response))
        else:
            kwargs = {}
            if quick is not None:   kwargs["quick_reply"] = quick
            if emojis:              kwargs["emojis"] = emojis      
            line.reply_message(event.reply_token, TextSendMessage(text=response, **kwargs))
        return

    def Push(self, line, chat_id, message):
        line.push_message(chat_id, message)  # Use Message Object; Prepare it before passing into this Function, that way it can cover all Message Types without clusters of if else
        return
    
    def PushImage(self, line, chat_id, image_url):
        message = ImageSendMessage(original_content_url=image_url, preview_image_url=image_url)
        line.push_message(chat_id, message)
        return
    
    def PushBoth(self, line, chat_id, response, image=False, quick=None, emojis=None):
        if image:   line.push_message(chat_id, ImageSendMessage(original_content_url=response, preview_image_url=response))
        else:
            kwargs = {}
            if quick is not None:   kwargs["quick_reply"] = quick
            if emojis:              kwargs["emojis"] = emojis
            line.push_message(chat_id, TextSendMessage(text=response, **kwargs))
        return

    def Save(self, chat_id, user_id, user_name, message_id, message, timestamp, content="Text", file_name=None, file_path=None, package_id=None, sticker_id=None):
        if content == "Text":
            message_document = {
                'user_id':      user_id,
                'user_name':    user_name,
                'message_id':   message_id,
                'message':      message,
                'content':      content,
                'timestamp':    timestamp
            }
        elif content == "Sticker":
            message_document = {
                'user_id':      user_id,
                'user_name':    user_name,
                'message_id':   message_id,
                'package_id':   file_name,
                'sticker_id':   file_path,
                'content':      content,
                'timestamp':    timestamp
            }
        else:
            #with open(file_path, "rb") as file_data:    file_id = self.Settings.MongoFS.put(file_data, filename=file_name)     # Store the media in GridFS
            message_document = {
                'user_id':      user_id,
                'user_name':    user_name,
                'message_id':   message_id,
                'file_path':    file_path,               # Change to file_id for Mongo
                'file_name':    file_name,
                'content':      content,
                'timestamp':    timestamp
            }

        # ⬆️ User save
        user_db = self.Settings.LineMessages.user_db(chat_id)                         # Per-user ledger
        user_db.insert_one(message_document)

        # ⬇️ Swap out Mongo with local JSONL
        self.Settings.LineMessages.insert_one("LineMessages", message_document)
        print(f"Message added to chat {chat_id} (Offline DB)\n")

        # MongoDB
        ## result = self.Settings.LineMessages.update_one({"chat_id": chat_id}, {"$push": {"messages": message_document}}, upsert=True); print(f"\nMessage added to chat {chat_id} with result: {result.modified_count}")
        return

    def Listener(self, *args):
        """
        Run in background. Whenever you type a single `m` + Enter, it will call multicast_prompt().
        """
        while True:
            line = sys.stdin.readline().strip()
            if line.lower() == "m":
                print("\n🔧 [stdin] 'm' detected → launching Multicast prompt…\n")
                #self.blueprint.Multicast(bTerminal=True)  NOTE I Don't think this is possible because you don't have a Class Instance of Greenprint
            # else: ignore other input

    def Weather(self, city):
        output = super().Weather(city)
        return output

# IG
class Relay(Interface):
    def __init__(self, settings):   super().__init__(settings)

    def Reply(self, chat_id, response):
        """
        Send a reply back to Instagram DM.
        Adapter (instagram.py) should implement the actual POST to Graph API,
        here we just define the interface shape.
        """
        print(f"[Relay] Reply → {chat_id}: {response}")
        return

    def Save(self, chat_id, user_id, user_name, message_id, message, timestamp, content="Text", **kwargs):
        """
        Persist Instagram message to DB.
        Mirrors Engine/Node/Hub style.
        """
        message_document = {
            'user_id':    user_id,
            'user_name':  user_name,
            'message_id': message_id,
            'message':    message,
            'content':    content,
            'timestamp':  timestamp
        }
        result = self.Collection.update_one(
            {"chat_id": chat_id},
            {"$push": {"messages": message_document}},
            upsert=True
        )
        print(f"[Relay] Message stored in {chat_id}, result: {result.modified_count}")
        return

    def Listener(self, *args):
        """
        Stub for any IG-specific listeners.
        For now, IG uses webhooks → handled in instagram.py.
        NOT NEEDED coz you control the Business Account
        """
        pass

    '''
    def Chat(
        self,
        event_or_update: Optional[Any],
        api: str,                    # "instagram"
        name: str,                   # e.g. self.character.FayName
        user_id: str,
        chat_id: str,
        prompt: str,
        *,
        message_id: Optional[str] = None,
        user_name: Optional[str] = None,
        timestamp: Optional[float] = None,
        attachments: Optional[List[Dict[str, Any]]] = None,
        typing_cb: Optional[Callable[[bool], None]] = None,   # transport typing indicator callback
        meta: Optional[Dict[str, Any]] = None,
    ) -> str:
        """
        1) (optional) typing_on
        2) persist user message
        3) generate with core brain (Interface hooks)
        4) send and persist assistant reply
        5) (optional) typing_off
        """
        t0 = time.time()
        if typing_cb:
            self._safe_call(lambda: typing_cb(True))

        # 1) Save inbound
        ts = timestamp or time.time()
        self.Save(
            chat_id=chat_id,
            user_id=user_id,
            user_name=user_name or "",
            message_id=message_id or f"ig:{int(ts*1000)}",
            message=prompt,
            timestamp=ts,
            content="Text",
            direction="in"
        )

        # 2) Build normalized convo + call core
        sys_prompt = self._system_prompt(name=name, api=api)
        convo = [
            Message("system", sys_prompt, {"api": api, "name": name}),
            Message("user", prompt, {"user_id": str(user_id), "chat_id": str(chat_id)}),
        ]
        reply = self._generate_reply(
            brain=self._route_model(api=api, meta=meta or {}),
            messages=convo,
            attachments=attachments or [],
        )
        reply = self._postprocess(reply, api=api, name=name)

        # 3) Send out
        self.Reply(chat_id, reply)

        # 4) Save outbound
        self.Save(
            chat_id=chat_id,
            user_id="fay",
            user_name=name,
            message_id=f"ig:bot:{int(time.time()*1000)}",
            message=reply,
            timestamp=time.time(),
            content="Text",
            direction="out"
        )

        # 5) telemetry
        ms = (time.time() - t0) * 1000
        log.info("[Relay] chat=%s user=%s ms=%.1f", chat_id, user_id, ms)

        if typing_cb:
            self._safe_call(lambda: typing_cb(False))
        return reply
    
    def _user_file(self, user_id: str, chat_id: str | None = None) -> str:
        """Return per-user JSONL path; nested by chat as optional second ledger."""
        uid  = self.Name(user_id)
        path = os.path.join(self.WechatMessages, uid)
        os.makedirs(path, exist_ok=True)
        # 1) primary per-user ledger
        user_ledger = os.path.join(path, f"{uid}.jsonl")
        # 2) optional per-chat ledger inside user folder
        if chat_id:
            cid  = self.Name(chat_id)
            chat_folder = os.path.join(path, "chats")
            os.makedirs(chat_folder, exist_ok=True)
            chat_ledger = os.path.join(chat_folder, f"{cid}.jsonl")
        else:  chat_ledger = None
        return user_ledger, chat_ledger

    def _append_jsonl(self, file_path: str, payload: dict):
        """Append a single JSON line (UTF-8) with a per-file lock."""
        if not file_path: return
        lock = self._locks[file_path]
        with lock:
            with open(file_path, "a", encoding="utf-8") as f:   f.write(json.dumps(payload, ensure_ascii=False) + "\n")
    '''

# Web
class Portal(Interface):
    def __init__(self, settings):   super().__init__(settings)
    
    def Save(self, chat_id, user_id, user_name, message, timestamp, content="Text", file_name=None, file_path=None, segment=None, nquestion=None):
        if content == "Text":
            message_document = {
                "user_id":   user_id,
                "user_name": user_name,
                "message":   message,
                "content":   content,
                "timestamp": timestamp
            }
        else:
            message_document = {
                "user_id":   user_id,
                "user_name": user_name,
                "file_name": file_name,
                "file_path": file_path,
                "content":   content,
                "timestamp": timestamp
            }

        # New
        if segment is not None:     message_document["segment"] = segment
        if nquestion is not None:
            try:                    message_document["nquestion"] = int(nquestion)
            except Exception:       message_document["nquestion"] = nquestion

        # Per-user JSONL
        user_db = self.Settings.WebMessages.scoped_db(user_id, "User")
        user_db.insert_one(message_document)

        chat_db = self.Settings.WebMessages.scoped_db(chat_id, "Chat")
        chat_db.insert_one(message_document)

        # Optional global log file
        self.Settings.WebMessages.insert_one("WebMessages", message_document)
        print(f"Message added to Web {user_name}: {chat_id}")   # Include Content ?

        # MongoDB
        ## result = self.Settings.WebMessages.update_one({"chat_id": chat_id}, {"$push": {"messages": message_document}}, upsert=True); print(f"\nMessage added to chat {chat_id} with result: {result.modified_count}")

# Mail
class Courier(Interface):
    def __init__(self, settings):   super().__init__(settings)

# Wechat
class Wall(Interface):
    def __init__(self, settings):   super().__init__(settings)

    def Reply(self, response): return

    def Save(self, chat_id, user_id, user_name, message_id, message, timestamp, content="Text", file_name=None, file_path=None, media_id=None):
        if content == "Text":
            message_document = {
                'user_id':      user_id,
                'user_name':    user_name,
                'chat_id':      chat_id,
                'message_id':   message_id,
                'message':      message,
                'content':      content,
                'timestamp':    timestamp
            }
        else:
            # with open(file_path, "rb") as file_data:    file_id = self.Settings.MongoFS.put(file_data, filename=file_name)     # Store the media in GridFS
            message_document = {
                'user_id':      user_id,
                'user_name':    user_name,
                'chat_id':      chat_id,
                'media_id':     media_id,
                'message_id':   message_id,
                'message':      message,
                'file_path':    file_path,                                              # Change to file_id for Mongo
                'file_name':    file_name,
                'content':      content,
                'timestamp':    timestamp
            }

        # ⬆️ User save
        user_db = self.Settings.WechatMessages.user_db(user_id)                         # Per-user ledger
        user_db.insert_one(message_document)

        # ⬇️ Global Save
        self.Settings.WechatMessages.insert_one("WechatMessages", message_document)     # Swapped out Mongo with local JSONL
        print(f"Message added to chat {chat_id} (Offline DB)\n")
        return

########################################################################
#########################          Web         #########################

class Article:
    def __init__(self, driver): self.driver = driver
    
    def process_driver(self):
        # What about Captcha ?
        # Needs to be Implicit; Wait 3 seconds for content to be loaded ? Who wrote this Comment Me or You ? 
        time.sleep(3)
        html = self.driver.page_source
        soup = BeautifulSoup(html, "html.parser")

        title = ""
        body = ""
        if hasattr(soup.title, 'string'): title = soup.title.string 
        
        for p_tags in soup.find_all('p'):
            body += p_tags.text
            body += "\n"
            
        # Close the webdriver only after we verified we got the data
        self.driver.close()

        nWords      = len(body.split())
        newsprompt  = (title or '') + body

        print(f'Number of Words:      {nWords}')
        print(f"Original Title:   {title}\n\n")

        return title, newsprompt
    
    def process_error(self, driver):
        raise InterfaceException("There is some error while processing this website.")
        # First we should check what error is it, login/popup error or some sort
        # Filter it either through a param from try except when called or handle the error here again
        # Then we perform clicks, login, signup, whatever that stops us in our way
        # NOTE Vincent Why this Class Have process_error but Youtube and Spotify Don't Have ? 

class Spotify:
    def __init__(self, driver): self.driver:  webdriver.Firefox  = driver

    def process_driver(self):
        try:
            time.sleep(2)
            wait = WebDriverWait(self.driver, timeout=5).until(lambda d : d.find_element(By.CLASS_NAME, "RP2rRchy4i8TIp1CTmb7"))
        except:
            print ("Cant find element")
            return ""
        
        html = self.driver.page_source
        soup = BeautifulSoup(html, "html.parser")
        border = soup.find('div', class_="RP2rRchy4i8TIp1CTmb7")
        song_name = border.find('h1', class_='ksSRyh')
        artist_name = border.find('a', attrs={'data-testid': 'creator-link'})
        return song_name.get_text(), artist_name.get_text()

class YoutubeMusic:
    def __init__(self, driver): self.driver: webdriver.Firefox = driver
    
    def process_driver(self):
        try:
            time.sleep(2)
            wait = WebDriverWait(self.driver, timeout=5).until(lambda d : d.find_element(By.TAG_NAME, "ytmusic-player-bar"))
        except:
            print ("Cant find element")
            return ""
        
        html            = self.driver.page_source
        soup            = BeautifulSoup(html, "html.parser")
        music_player    = soup.find('ytmusic-player-bar')
        song_name       = music_player.find('yt-formatted-string', class_='title style-scope ytmusic-player-bar')
        artist_name     = music_player.find('a', 'yt-simple-endpoint style-scope yt-formatted-string')
        return song_name.get_text(), artist_name.get_text()

class YTDLSource(PCMVolumeTransformer):
    # Class methods:    Setup yt-dlp options [ Used in MusicStream { Engine(Interface) } ]
    ytdl_format_options = {
        'format': 'bestaudio/best',
        'outtmpl': '%(extractor)s-%(id)s-%(title)s.%(ext)s',
        'restrictfilenames': True,
        'noplaylist': True,
        'nocheckcertificate': True,
        'ignoreerrors': False,
        'logtostderr': False,
        'quiet': True,
        'no_warnings': True,
        'default_search': 'auto',
        'source_address': '0.0.0.0'  # Bind to ipv4 since ipv6 addresses cause issues sometimes
    }

    ffmpeg_options      = {  'options': '-vn'   }

    ytdl                = yt_dlp.YoutubeDL(ytdl_format_options)

    def __init__(self, source, data, volume=0.5):
        super().__init__(source, volume)

        self.data   = data
        self.title  = data.get('title')
        self.url    = data.get('url')
        
    @classmethod
    async def from_url(cls, url, *, loop=None, stream=False):
        loop = loop or asyncio.get_event_loop()
        data = await loop.run_in_executor(None, lambda: YTDLSource.ytdl.extract_info(url, download=not stream))

        if 'entries' in data:   data = data['entries'][0]   # First item
        filename = data['url'] if stream else YTDLSource.ytdl.prepare_filename(data)
        return (cls(FFmpegPCMAudio(filename, **YTDLSource.ffmpeg_options), data=data), filename)

class Maps:
    def __init__(self, driver): self.driver = driver

    def process_driver(self, url, html=""):
        final_url = url or ""
        title     = (getattr(self.driver, "title", "") or "").strip() if self.driver else ""
        label     = self._clean_label(title)

        lat, lon  = self._coords_from_url(final_url)
        if (lat is None or lon is None) and html:
            lat, lon = self._coords_from_html(html)

        # IMPORTANT: define a short query string (never a URL)
        if lat is not None and lon is not None:
            query = f"{lat},{lon}"
        elif label:
            query = label
        else:
            query = ""  # last resort

        # Clean "open" link
        open_url = (
            f"https://www.google.com/maps/search/?api=1&query={urllib.parse.quote(query)}"
            if query else final_url
        )

        return {
            "type": "map",
            "lat": lat,
            "lon": lon,
            "label": label,
            "query": query,          # <- THIS is what Line.py should pair on
            "open_url": open_url,    # <- always clean
            "source_url": final_url,
        }

    def _provider_from_url(self, url: str) -> str:
        u = (url or "").lower()
        if "google.com/maps" in u or "maps.app.goo.gl" in u or "goo.gl/maps" in u: return "google"
        if "maps.apple.com" in u: return "apple"
        if "openstreetmap.org" in u: return "osm"
        if "waze.com" in u: return "waze"
        if "here.com" in u or "wego.here.com" in u: return "here"
        return "map"

    def _clean_label(self, title: str) -> str:
        if not title: return ""
        for suffix in (
            " - Google Maps",
            " - Apple Maps",
            " - OpenStreetMap",
            " | Waze",
            " - Waze",
            " - HERE WeGo",
            " - HERE",
        ):
            if title.endswith(suffix):
                title = title[:-len(suffix)].strip()

        low = title.lower()
        if low in ("google maps", "maps", "apple maps", "openstreetmap", "waze", "here wego", "here"):
            return ""
        return title

    def _coords_from_url(self, url: str):
        try:
            u = urllib.parse.unquote(url or "")
            p = urllib.parse.urlparse(u)
            qs = urllib.parse.parse_qs(p.query)

            for key in ("q", "query", "ll", "center", "daddr", "saddr", "destination", "origin", "cp", "where"):
                if key in qs and qs[key]:
                    cand = qs[key][0]
                    lat, lon = self._coords_from_text(cand)
                    if self._valid(lat, lon):
                        return lat, lon

            if "mlat" in qs and "mlon" in qs:
                lat = self._to_float(qs.get("mlat", [""])[0])
                lon = self._to_float(qs.get("mlon", [""])[0])
                if self._valid(lat, lon):
                    return lat, lon

            m = re.search(r"@(-?\d{1,2}\.\d+),\s*(-?\d{1,3}\.\d+)", u)
            if m:
                lat = self._to_float(m.group(1))
                lon = self._to_float(m.group(2))
                if self._valid(lat, lon):
                    return lat, lon

            m = re.search(r"!3d(-?\d+(?:\.\d+)?)!4d(-?\d+(?:\.\d+)?)", u)
            if m:
                lat = self._to_float(m.group(1))
                lon = self._to_float(m.group(2))
                if self._valid(lat, lon):
                    return lat, lon

            if p.fragment:
                m = re.search(r"map=\d+\/(-?\d{1,2}\.\d+)\/(-?\d{1,3}\.\d+)", p.fragment)
                if m:
                    lat = self._to_float(m.group(1))
                    lon = self._to_float(m.group(2))
                    if self._valid(lat, lon):
                        return lat, lon

            if p.scheme == "geo":
                lat, lon = self._coords_from_text(u)
                if self._valid(lat, lon):
                    return lat, lon

        except Exception as e:
            print(e)

        return None, None

    def _coords_from_html(self, html: str):
        try:
            # Explicit lat/lng patterns only (safe)
            m = re.search(
                r"\b(?:lat|latitude)\b\s*[:=]\s*(-?\d{1,2}\.\d+).{0,80}?\b(?:lng|lon|longitude)\b\s*[:=]\s*(-?\d{1,3}\.\d+)",
                html, re.I | re.S
            )
            if m:
                lat = self._to_float(m.group(1))
                lon = self._to_float(m.group(2))
                if self._valid(lat, lon):
                    return lat, lon

            # Google '@lat,lon'
            m = re.search(r"@(-?\d{1,2}\.\d+),\s*(-?\d{1,3}\.\d+)", html)
            if m:
                lat = self._to_float(m.group(1))
                lon = self._to_float(m.group(2))
                if self._valid(lat, lon):
                    return lat, lon

            # Google '!3dLAT!4dLON'
            m = re.search(r"!3d(-?\d+(?:\.\d+)?)!4d(-?\d+(?:\.\d+)?)", html)
            if m:
                lat = self._to_float(m.group(1))
                lon = self._to_float(m.group(2))
                if self._valid(lat, lon):
                    return lat, lon

        except Exception as e:
            print(e)

        return None, None

    def _coords_from_text(self, s: str):
        if not s: return None, None
        t = urllib.parse.unquote(str(s))
        t = re.sub(r"^(loc:|geo:)", "", t.strip(), flags=re.I)

        m = re.search(r"(-?\d{1,2}\.\d+)\s*,\s*(-?\d{1,3}\.\d+)", t)
        if not m: return None, None
        return self._to_float(m.group(1)), self._to_float(m.group(2))

    def _to_float(self, x):
        try: return float(str(x).strip())
        except: return None

    def _valid(self, lat, lon):
        return lat is not None and lon is not None and abs(lat) <= 90 and abs(lon) <= 180

    @classmethod
    def Extract(cls, body: str):
        # optional legacy helper
        if not body or not isinstance(body, str): return None
        if "Open in Maps" not in body: return None

        m = re.search(r"Open in Maps\s*→\s*(https?://\S+)", body)
        if not m: return None
        open_url = m.group(1).strip()

        label = ""
        first = body.splitlines()[0].strip() if body.splitlines() else ""
        if first.startswith("📍"):
            label = first.replace("📍", "", 1).strip()
            if label.lower() in ("got it!", "got it", ""):
                label = ""

        tmp = cls(None)
        lat, lon = tmp._coords_from_url(open_url)
        return {"lat": lat, "lon": lon, "label": label, "open_url": open_url}

########################################################################
#######################           Error          #######################

class InterfaceException(Exception):
    def __init__(self, message): self.msg = message

    def __repr__(self): return self.msg
    
    def __str__(self):  return self.msg

########################################################################